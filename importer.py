# =============================================================================
# Policy Coherence Kit -- importer.py
# Full matrix import from multi-sheet Excel workbooks.
# Pure logic, no tkinter.
#
# Public API
# ----------
# import_matrices_from_excel(path) -> ImportResult
#
# ImportResult
# ------------
# .matrices  : List[PolicyMatrix]   -- one per sheet, in sheet order
# .warnings  : List[str]            -- one entry per incomplete matrix
# =============================================================================

from dataclasses import dataclass, field
from typing import List, Tuple

from models import PolicyMatrix, generate_codes
from constants import COHERENCE_RATINGS, DIAGONAL_VALUE


@dataclass
class ImportResult:
    """
    Result of a full-workbook import.

    matrices : successfully parsed PolicyMatrix objects (may be incomplete)
    warnings : human-readable messages for sheets with blank cells
    """
    matrices: List[PolicyMatrix] = field(default_factory=list)
    warnings: List[str]          = field(default_factory=list)


# =============================================================================
# Main entry point
# =============================================================================

def import_matrices_from_excel(path: str) -> ImportResult:
    """
    Read every sheet in the workbook as one decision-maker matrix.

    Expected sheet layout
    ---------------------
    Row 1, Col 1  : ignored (corner cell)
    Row 1, Col 2+ : full policy names (column headers)
    Col 1, Row 2+ : full policy names (row headers, must match col headers)
    Data cells    : one of the 7 coherence rating labels, or blank

    The sheet name becomes the decision-maker name.
    Policy names are auto-coded P1, P2, ...

    Raises
    ------
    ImportError  if openpyxl is not installed
    ValueError   if a sheet has fewer than 2 policies, mismatched headers,
                 or unrecognised rating values
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel import.\n"
            "Install it with:  pip install openpyxl"
        )

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result = ImportResult()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            result.warnings.append(
                f'Sheet "{sheet_name}": empty, skipped.'
            )
            continue

        # ---- Read column headers (row 0, cols 1+) ----
        col_headers = [
            str(v).strip() for v in rows[0][1:]
            if v is not None and str(v).strip()
        ]

        if len(col_headers) < 2:
            result.warnings.append(
                f'Sheet "{sheet_name}": fewer than 2 policies in header row, skipped.'
            )
            continue

        n = len(col_headers)

        # ---- Read row headers (col 0, rows 1+) ----
        row_headers = []
        for row in rows[1:n + 1]:
            val = row[0]
            row_headers.append(str(val).strip() if val is not None else "")

        # ---- Validate headers match ----
        if row_headers != col_headers:
            # Try to give a helpful message
            mismatches = [
                f'  row {i+1}: "{row_headers[i]}" vs col "{col_headers[i]}"'
                for i in range(min(len(row_headers), len(col_headers)))
                if i < len(row_headers) and row_headers[i] != col_headers[i]
            ]
            detail = "\n".join(mismatches[:5])
            raise ValueError(
                f'Sheet "{sheet_name}": row and column headers do not match.\n'
                f'{detail}'
            )

        policies = col_headers
        codes    = generate_codes(policies)
        matrix   = PolicyMatrix(
            decision_maker=sheet_name,
            policies=policies,
            codes=codes,
        )

        # ---- Read data cells ----
        blanks        = []
        bad_values    = []

        for i in range(n):
            data_row_idx = i + 1          # row in the sheet (0-based)
            if data_row_idx >= len(rows):
                # Row missing entirely — all cells blank
                for j in range(n):
                    if i != j:
                        blanks.append((codes[i], codes[j]))
                continue

            row_data = rows[data_row_idx]

            for j in range(n):
                col_idx = j + 1           # col in the sheet (0-based)
                if i == j:
                    continue              # diagonal always Neutral

                raw = row_data[col_idx] if col_idx < len(row_data) else None
                val = str(raw).strip() if raw is not None else ""

                if not val or val.lower() == "none":
                    blanks.append((codes[i], codes[j]))
                    continue

                # Normalise capitalisation to match our vocabulary
                matched = _match_rating(val)
                if matched is None:
                    bad_values.append(
                        f'  {codes[i]}->{codes[j]}: "{val}"'
                    )
                else:
                    matrix.set_rating(i, j, matched)

        # ---- Report bad values as an error ----
        if bad_values:
            raise ValueError(
                f'Sheet "{sheet_name}": unrecognised rating values:\n'
                + "\n".join(bad_values[:10])
                + ("\n  ..." if len(bad_values) > 10 else "")
                + f'\n\nAllowed values: {", ".join(COHERENCE_RATINGS)}'
            )

        # ---- Record blanks as a warning (not an error) ----
        if blanks:
            cells = ", ".join(f"{r}->{c}" for r, c in blanks[:10])
            suffix = f" (+{len(blanks)-10} more)" if len(blanks) > 10 else ""
            result.warnings.append(
                f'"{sheet_name}": {len(blanks)} blank cell(s): {cells}{suffix}'
            )

        result.matrices.append(matrix)

    wb.close()
    return result


# =============================================================================
# Helpers
# =============================================================================

_RATINGS_LOWER = {r.lower(): r for r in COHERENCE_RATINGS}

def _match_rating(value: str):
    """
    Case-insensitive match against the 7 coherence rating labels.
    Returns the canonical label or None if no match.
    """
    return _RATINGS_LOWER.get(value.lower())
