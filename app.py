# =============================================================================
# Policy Coherence Kit -- app.py
# Multi-project workspace. Each project lives in its own top-level notebook tab
# and has completely independent state: matrices, analysis results, etc.
# =============================================================================

import io
import math
import sys
from pathlib import Path
import tkinter as tk
from PIL import Image, ImageDraw, ImageTk
from tkinter import ttk, messagebox, filedialog
import tkinter.font as tkFont
from typing import List, Optional
from dataclasses import dataclass, field

from models import PolicyMatrix, make_empty_matrix
from matrix_widget import MatrixWidget
from dialogs import NewMatrixDialog, ProjectSetupDialog, _SimpleInputDialog, style_button
from aggregator import (
    check_completeness,
    aggregate_average, aggregate_majority, aggregate_weighted, resolve_ties,
    AggregationResult,
)
from aggregation_dialog import (
    AggregationMethodDialog, WeightDialog, TieResolutionDialog,
)
from aggregation_tab import AggregationTab
from coherence_scores_tab import CoherenceScoresTab
from range_of_influence_tab import RangeOfInfluenceTab
from pca_tab import PCATab
from network_tab import NetworkTab
from llm_tab import LLMInterpretationTab
from results_insights_tab import ResultsInsightsTab
from importer import import_matrices_from_excel
from constants import (
    APP_TITLE, WINDOW_MIN_WIDTH, WINDOW_MIN_HEIGHT,
    FONT_FAMILY, FONT_SIZE_SMALL, FONT_SIZE_NORMAL,
    FONT_SIZE_HEADER, FONT_SIZE_TITLE,
    COLOR_BG, COLOR_PANEL, COLOR_ACCENT, COLOR_ACCENT2,
    COLOR_TEXT, COLOR_TEXT_LIGHT, COLOR_BORDER,
    COLOR_TAB_BG, COLOR_BUTTON, COLOR_BUTTON_FG,
    CURSOR_HAND,
)


def _hex_interp(c1, c2, t):
    """Interpolate between two hex colours. t=0 → c1, t=1 → c2."""
    t = max(0.0, min(1.0, t))
    r = int(int(c1[1:3], 16) + (int(c2[1:3], 16) - int(c1[1:3], 16)) * t)
    g = int(int(c1[3:5], 16) + (int(c2[3:5], 16) - int(c1[3:5], 16)) * t)
    b = int(int(c1[5:7], 16) + (int(c2[5:7], 16) - int(c1[5:7], 16)) * t)
    return f"#{r:02x}{g:02x}{b:02x}"


def _rrect_pts(x1, y1, x2, y2, r, steps=10):
    """Return polygon points for a rounded rectangle."""
    pts = []
    for cx, cy, a0 in [(x2-r, y1+r, -90), (x2-r, y2-r, 0),
                        (x1+r, y2-r,  90), (x1+r, y1+r, 180)]:
        for i in range(steps + 1):
            a = math.radians(a0 + 90 * i / steps)
            pts += [cx + r * math.cos(a), cy + r * math.sin(a)]
    return pts


def _scroll_canvas_x_by_pixels(canvas, dx_px):
    """Move a horizontally scrollable canvas by a pixel delta."""
    x1, x2 = canvas.xview()
    span = x2 - x1
    if span >= 1.0:
        return False
    cw = canvas.winfo_width()
    if cw <= 0:
        return False
    next_x = max(0.0, min(1.0 - span, x1 + dx_px * span / cw))
    if abs(next_x - x1) < 1e-6:
        return False
    canvas.xview_moveto(next_x)
    return True


def _animate_canvas_x_scroll(canvas, state, delta_px, duration=180):
    """Animate horizontal canvas scrolling with eased motion."""
    if not delta_px:
        return
    if state["anim"][0]:
        canvas.after_cancel(state["anim"][0])
        state["anim"][0] = None

    steps = max(8, int(duration / 16))
    remaining = [float(delta_px)]
    frame = [0]

    def _tick():
        if frame[0] >= steps:
            state["anim"][0] = None
            return
        frame[0] += 1
        portion = 1.0 / (steps - frame[0] + 1)
        move_px = remaining[0] * portion
        remaining[0] -= move_px
        if _scroll_canvas_x_by_pixels(canvas, move_px):
            state["anim"][0] = canvas.after(16, _tick)
        else:
            state["anim"][0] = None

    state["anim"][0] = canvas.after(16, _tick)


# =============================================================================
# Project data container
# =============================================================================

@dataclass
class _FlowFrame(tk.Frame):
    """Horizontal-wrapping frame — like `display:flex; flex-wrap:wrap` in CSS.
    Add children with .add(widget, padx=N) instead of packing them directly."""

    def __init__(self, *args, vgap=6, **kwargs):
        super().__init__(*args, **kwargs)
        self._items = []
        self._vgap  = vgap
        self.bind("<Configure>", self._reflow)

    def add(self, widget, padx=0, is_sep=False):
        self._items.append((widget, padx, is_sep))

    def _reflow(self, e=None):
        W = self.winfo_width()
        if W <= 1 or not self._items:
            return
        # Pass 1: group into rows
        rows = []
        row = []; row_w = 0
        for widget, padx, is_sep in self._items:
            widget.update_idletasks()
            req_w = widget.winfo_reqwidth() + padx * 2
            req_h = widget.winfo_reqheight()
            if row_w > 0 and row_w + req_w > W:
                rows.append(row); row = []; row_w = 0
            row.append((widget, padx, is_sep, req_w, req_h))
            row_w += req_w
        if row:
            rows.append(row)
        # Pass 2: strip separators from row edges and hide them
        hidden = set()
        clean  = []
        for row in rows:
            while row and row[0][2]:  hidden.add(row[0][0]); row = row[1:]
            while row and row[-1][2]: hidden.add(row[-1][0]); row = row[:-1]
            if row:
                clean.append(row)
        for w in hidden:
            w.place_forget()
        # Pass 3: place each row right-aligned
        y = 0
        for row in clean:
            row_total = sum(rw for _, _, _, rw, _ in row)
            row_h     = max(rh for _, _, _, _, rh in row)
            x = W - row_total
            for widget, padx, _, req_w, _ in row:
                widget.place(x=x + padx, y=y)
                x += req_w
            y += row_h + self._vgap
        new_h = y - self._vgap if clean else 0
        if self.winfo_reqheight() != new_h:
            self.configure(height=new_h)


@dataclass
class Project:
    """All state belonging to one project."""
    name:        str
    matrices:    List[PolicyMatrix] = field(default_factory=list)
    agg_results: List[AggregationResult] = field(default_factory=list)
    agg_tab_ids: List[str] = field(default_factory=list)   # inner notebook tab ids
    analysis_tab_meta: dict = field(default_factory=dict, repr=False)
    notebook:    Optional[ttk.Notebook] = field(default=None, repr=False)
    frame:       Optional[tk.Frame]     = field(default=None, repr=False)


# =============================================================================
# Project name dialog (shown on launch + new project)
# =============================================================================

class _ProjectNameDialog(tk.Toplevel):
    """Simple modal to collect a project name."""

    def __init__(self, parent, title="New Project", existing_names=None):
        super().__init__(parent)
        self.transient(parent)
        self.title(title)
        self.configure(bg=COLOR_BG)
        self.resizable(False, False)
        self.grab_set()
        self._existing = existing_names or []
        self.result: Optional[str] = None
        self._build()
        self.update_idletasks()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        w, h   = self.winfo_reqwidth(), self.winfo_reqheight()
        self.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

    def _build(self):
        tk.Label(
            self, text="Project Name",
            font=(FONT_FAMILY, FONT_SIZE_TITLE, "bold"),
            bg=COLOR_BG, fg=COLOR_ACCENT,
        ).pack(anchor="w", padx=24, pady=(20, 4))
        tk.Label(
            self,
            text="Give this project a unique name.",
            font=(FONT_FAMILY, FONT_SIZE_NORMAL, "italic"),
            bg=COLOR_BG, fg=COLOR_TEXT_LIGHT,
        ).pack(anchor="w", padx=24, pady=(0, 10))
        ttk.Separator(self).pack(fill="x", padx=24, pady=4)
        self._var = tk.StringVar()
        self._entry = tk.Entry(
            self, textvariable=self._var,
            font=(FONT_FAMILY, FONT_SIZE_NORMAL),
            bg="#fafbfc", fg=COLOR_TEXT,
            insertbackground=COLOR_ACCENT,
            relief="solid", bd=1, width=36,
        )
        self._entry.pack(padx=24, pady=(8, 16))
        self._entry.focus_set()
        ttk.Separator(self).pack(fill="x", padx=24, pady=4)
        btn_row = tk.Frame(self, bg=COLOR_BG)
        btn_row.pack(anchor="e", padx=24, pady=(6, 18))
        ok = tk.Button(btn_row, text="Create Project", command=self._on_ok)
        style_button(ok)
        ok.pack()
        self.bind("<Return>", lambda e: self._on_ok())
        self.bind("<Escape>", lambda e: self.destroy())

    def _on_ok(self):
        name = self._var.get().strip()
        if not name:
            messagebox.showwarning("Name Required",
                                   "Please enter a project name.", parent=self)
            return
        if name in self._existing:
            messagebox.showwarning("Duplicate Name",
                                   f'A project named "{name}" already exists.',
                                   parent=self)
            return
        self.result = name
        self.destroy()


# =============================================================================
# Main application
# =============================================================================

class PolicyCoherenceApp:
    """
    Multi-project workspace controller.

    Top-level notebook  : one tab per project
    Inner notebook      : DM matrices + analysis tabs for that project
    """

    def __init__(self, root: tk.Tk):
        self.root     = root
        self.projects: List[Project] = []

        self._configure_root()
        self._apply_styles()
        self._build_ui()


    # ==================================================================
    # Setup
    # ==================================================================

    def _configure_root(self):
        self.root.title(APP_TITLE)
        self.root.minsize(WINDOW_MIN_WIDTH, WINDOW_MIN_HEIGHT)
        self.root.configure(bg=COLOR_BG)
        self.root.update_idletasks()
        w, h = 1200, 780
        sw, sh = self.root.winfo_screenwidth(), self.root.winfo_screenheight()
        self.root.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

    def _apply_styles(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TNotebook", background=COLOR_BG,
                        borderwidth=0, relief="flat", padding=0,
                        tabmargins=[0, 0, 0, 0],
                        bordercolor=COLOR_BG, darkcolor=COLOR_BG, lightcolor=COLOR_BG)
        style.layout("TNotebook", [("Notebook.client", {"sticky": "nswe"})])
        style.configure("TNotebook.Tab", background=COLOR_TAB_BG,
                        foreground=COLOR_TEXT, padding=[14, 6],
                        font=(FONT_FAMILY, FONT_SIZE_NORMAL))
        style.map("TNotebook.Tab",
                  background=[("selected", COLOR_ACCENT)],
                  foreground=[("selected", "#ffffff")],
                  expand=[("selected", [1, 1, 1, 0])])
        style.configure("TScrollbar", background=COLOR_PANEL,
                        troughcolor=COLOR_BG, borderwidth=0, arrowsize=12)
        style.configure("TCombobox", fieldbackground="#ffffff",
                        background="#ffffff",
                        selectbackground=COLOR_ACCENT,
                        selectforeground="#ffffff")
        style.configure("TSeparator", background=COLOR_BORDER)

        # Outer project notebook — tabs hidden (switching done via sidebar)
        style.configure("Headless.TNotebook", background=COLOR_BG,
                        borderwidth=0, relief="flat", padding=0,
                        tabmargins=[0, 0, 0, 0],
                        bordercolor=COLOR_BG, darkcolor=COLOR_BG, lightcolor=COLOR_BG)
        style.layout("Headless.TNotebook",     [("Notebook.client", {"sticky": "nswe"})])
        style.layout("Headless.TNotebook.Tab", [])   # no tab rendering

        # Inner notebook style (slightly different tab colour)
        style.configure("Inner.TNotebook", background=COLOR_BG,
                        borderwidth=0, relief="flat", padding=0,
                        tabmargins=[0, 0, 0, 0],
                        bordercolor=COLOR_BG, darkcolor=COLOR_BG, lightcolor=COLOR_BG)
        style.layout("Inner.TNotebook", [("Notebook.client", {"sticky": "nswe"})])
        style.configure("Inner.TNotebook.Tab",
                        background="#c8c4bc",
                        foreground=COLOR_TEXT,
                        padding=[10, 4],
                        font=(FONT_FAMILY, FONT_SIZE_SMALL))
        style.map("Inner.TNotebook.Tab",
                  background=[("selected", COLOR_ACCENT2)],
                  foreground=[("selected", "#ffffff")],
                  expand=[("selected", [1, 1, 1, 0])])

    # ==================================================================
    # Build UI
    # ==================================================================

    def _build_ui(self):
        outer = tk.Frame(self.root, bg=COLOR_BG)
        outer.pack(fill="both", expand=True)

        self._build_sidebar(outer)

        self._content = tk.Frame(outer, bg=COLOR_BG)
        self._content.pack(side="left", fill="both", expand=True)

        self._build_topbar()
        self._build_statusbar()
        self._build_project_notebook()
        self._proj_nb.bind("<<NotebookTabChanged>>", self._on_project_tab_changed)
        self._update_topbar()

    def _build_sidebar(self, parent):
        self._sidebar_open_states = {}
        self._sidebar_expanded    = True
        _EXP_W = 320
        _COL_W = 80

        sidebar = tk.Frame(parent, bg="#fafbfc", width=_EXP_W)
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)

        # Right border
        tk.Frame(parent, bg="#e6e6e6", width=1).pack(side="left", fill="y")
        self._sidebar       = sidebar
        self._sidebar_exp_w = _EXP_W
        self._sidebar_col_w = _COL_W

        # ── Expanded header ────────────────────────────────────────────
        hdr_exp = tk.Frame(sidebar, bg="#fafbfc", height=90, cursor=CURSOR_HAND)
        hdr_exp.pack(fill="x")
        hdr_exp.pack_propagate(False)
        self._sidebar_hdr_exp = hdr_exp

        hdr_row = tk.Frame(hdr_exp, bg="#fafbfc", cursor=CURSOR_HAND)
        hdr_row.place(relx=0, rely=0.5, anchor="w", x=20)

        # Waypoints icon with badge
        _LOGO_ICON   = 17           # icon drawing area (px, within badge)
        _LOGO_PAD    = 9            # padding around icon inside badge
        _BADGE_SIZE  = _LOGO_ICON + _LOGO_PAD * 2   # = 35
        _BADGE_R     = 5
        _BADGE_BG    = "#1f2937"
        _LOGO_COLOR  = "#f5f7fa"
        logo_c = tk.Canvas(hdr_row, width=_BADGE_SIZE+2, height=_BADGE_SIZE+2,
                           bg="#fafbfc", highlightthickness=0, cursor=CURSOR_HAND)
        logo_c.pack(side="left", padx=(0, 10))

        def _draw_blocks_icon():
            B  = _BADGE_SIZE
            r  = _BADGE_R
            d  = r * 2
            # Badge background (arc+rect rounded square)
            bkw = dict(fill=_BADGE_BG, outline=_BADGE_BG)
            logo_c.create_arc(1,     1,     1+d,   1+d,   start=90,  extent=90, **bkw)
            logo_c.create_arc(B+1-d, 1,     B+1,   1+d,   start=0,   extent=90, **bkw)
            logo_c.create_arc(1,     B+1-d, 1+d,   B+1,   start=180, extent=90, **bkw)
            logo_c.create_arc(B+1-d, B+1-d, B+1,   B+1,   start=270, extent=90, **bkw)
            logo_c.create_rectangle(1+r, 1,   B+1-r, B+1,   fill=_BADGE_BG, outline=_BADGE_BG)
            logo_c.create_rectangle(1,   1+r, B+1,   B+1-r, fill=_BADGE_BG, outline=_BADGE_BG)
            # Waypoints icon (scaled to _LOGO_ICON, offset by pad+1)
            s  = _LOGO_ICON / 24.0
            o  = _LOGO_PAD + 1      # badge padding + 1px canvas inset
            lw = 1.35
            lkw = dict(fill=_LOGO_COLOR, width=lw, capstyle="round", joinstyle="round")
            ckw = dict(outline=_LOGO_COLOR, fill="", width=lw)
            logo_c.create_line(10.586*s+o, 5.414*s+o,  5.414*s+o, 10.586*s+o, **lkw)
            logo_c.create_line(18.586*s+o, 13.414*s+o, 13.414*s+o, 18.586*s+o, **lkw)
            logo_c.create_line(6*s+o, 12*s+o, 18*s+o, 12*s+o, **lkw)
            cr = 2*s
            for cx, cy in [(12,4), (12,20), (4,12), (20,12)]:
                logo_c.create_oval(cx*s+o-cr, cy*s+o-cr, cx*s+o+cr, cy*s+o+cr, **ckw)

        _draw_blocks_icon()

        text_col = tk.Frame(hdr_row, bg="#fafbfc", cursor=CURSOR_HAND)
        text_col.pack(side="left")

        _title_lbl = tk.Label(
            text_col, text="Policy Coherence Kit",
            font=(FONT_FAMILY, 20, "bold"),
            bg="#fafbfc", fg="#1f2937", justify="left", cursor=CURSOR_HAND,
        )
        _title_lbl.pack(anchor="w")

        _slogan_lbl = tk.Label(
            text_col,
            text="Evaluate interactions between policies using multiple decision-makers",
            font=(FONT_FAMILY, 10),
            bg="#fafbfc", fg="#a3a3a3", justify="left",
            wraplength=220, cursor=CURSOR_HAND,
        )
        _slogan_lbl.pack(anchor="w")

        for _w in (hdr_exp, hdr_row, logo_c, text_col, _title_lbl, _slogan_lbl):
            _w.bind("<Button-1>", lambda e: self._toggle_sidebar())

        # ── Collapsed header ───────────────────────────────────────────
        hdr_col = tk.Frame(sidebar, bg="#fafbfc", height=90, cursor=CURSOR_HAND)
        hdr_col.pack_propagate(False)
        self._sidebar_hdr_col = hdr_col

        col_logo_c = tk.Canvas(hdr_col, width=_BADGE_SIZE+2, height=_BADGE_SIZE+2,
                               bg="#fafbfc", highlightthickness=0, cursor=CURSOR_HAND)
        col_logo_c.place(relx=0.5, rely=0.5, anchor="center")

        def _draw_col_badge():
            B = _BADGE_SIZE; r = _BADGE_R; d = r * 2
            bkw = dict(fill=_BADGE_BG, outline=_BADGE_BG)
            col_logo_c.create_arc(1,     1,     1+d,   1+d,   start=90,  extent=90, **bkw)
            col_logo_c.create_arc(B+1-d, 1,     B+1,   1+d,   start=0,   extent=90, **bkw)
            col_logo_c.create_arc(1,     B+1-d, 1+d,   B+1,   start=180, extent=90, **bkw)
            col_logo_c.create_arc(B+1-d, B+1-d, B+1,   B+1,   start=270, extent=90, **bkw)
            col_logo_c.create_rectangle(1+r, 1,   B+1-r, B+1,   fill=_BADGE_BG, outline=_BADGE_BG)
            col_logo_c.create_rectangle(1,   1+r, B+1,   B+1-r, fill=_BADGE_BG, outline=_BADGE_BG)
            s  = _LOGO_ICON / 24.0
            o  = _LOGO_PAD + 1
            lw = 1.35
            lkw = dict(fill=_LOGO_COLOR, width=lw, capstyle="round", joinstyle="round")
            ckw = dict(outline=_LOGO_COLOR, fill="", width=lw)
            col_logo_c.create_line(10.586*s+o, 5.414*s+o,  5.414*s+o, 10.586*s+o, **lkw)
            col_logo_c.create_line(18.586*s+o, 13.414*s+o, 13.414*s+o, 18.586*s+o, **lkw)
            col_logo_c.create_line(6*s+o, 12*s+o, 18*s+o, 12*s+o, **lkw)
            cr = 2*s
            for cx, cy in [(12,4), (12,20), (4,12), (20,12)]:
                col_logo_c.create_oval(cx*s+o-cr, cy*s+o-cr, cx*s+o+cr, cy*s+o+cr, **ckw)

        _draw_col_badge()
        hdr_col.bind("<Button-1>", lambda e: self._toggle_sidebar())
        col_logo_c.bind("<Button-1>", lambda e: self._toggle_sidebar())

        # ── Divider + PROJECTS label (always in layout for spacing) ──────
        exp_content = tk.Frame(sidebar, bg="#fafbfc")
        exp_content.pack(fill="x")
        self._sidebar_exp_content = exp_content

        tk.Frame(exp_content, bg="#e6e6e6", height=1).pack(fill="x", pady=(0, 4))
        self._sidebar_proj_lbl = tk.Label(
            exp_content, text="PROJECTS",
            font=(FONT_FAMILY, FONT_SIZE_SMALL, "bold"),
            bg="#fafbfc", fg="#a3a3a3", anchor="w",
        )
        self._sidebar_proj_lbl.pack(fill="x", padx=16, pady=(12, 10))

        # ── Project list — scrollable container ───────────────────────
        _proj_scroll_outer = tk.Frame(sidebar, bg="#fafbfc")
        _proj_scroll_outer.pack(fill="both", expand=True)

        _proj_vbar = tk.Scrollbar(_proj_scroll_outer, orient="vertical", width=5)
        _proj_vbar.pack(side="right", fill="y")

        _proj_canvas = tk.Canvas(_proj_scroll_outer, bg="#fafbfc",
                                 highlightthickness=0,
                                 yscrollcommand=_proj_vbar.set,
                                 yscrollincrement=20)
        _proj_canvas.pack(side="left", fill="both", expand=True)
        _proj_vbar.configure(command=_proj_canvas.yview)

        _proj_inner = tk.Frame(_proj_canvas, bg="#fafbfc")
        _proj_win   = _proj_canvas.create_window((0, 0), window=_proj_inner,
                                                  anchor="nw")

        _proj_canvas.bind("<Configure>",
            lambda e: _proj_canvas.itemconfig(_proj_win, width=e.width))
        _proj_inner.bind("<Configure>",
            lambda e: _proj_canvas.configure(
                scrollregion=_proj_canvas.bbox("all")))

        _proj_canvas.bind("<MouseWheel>",
            lambda e: _proj_canvas.yview_scroll(-1 if e.delta > 0 else 1, "units"))
        _proj_canvas.bind("<Button-4>",
            lambda e: _proj_canvas.yview_scroll(-1, "units"))
        _proj_canvas.bind("<Button-5>",
            lambda e: _proj_canvas.yview_scroll( 1, "units"))

        self._sidebar_proj_list   = _proj_inner
        self._sidebar_proj_canvas = _proj_canvas

        # Shared New Project button constants (used by both collapsed + expanded)
        _NP_BG     = "#eaeef4"
        _NP_HOVER  = "#d0dbe7"
        _NP_FG     = "#1f2937"
        _NP_RADIUS = 5
        _NP_H      = 40
        _NP_ICON   = 18
        _NP_GAP    = 6
        _NP_LABEL  = "New Project"

        _np_font   = tkFont.Font(family=FONT_FAMILY, size=13)
        _np_text_w = _np_font.measure(_NP_LABEL)
        _np_block  = _NP_ICON + _NP_GAP + _np_text_w   # total icon+gap+text width

        def _draw_np_icon(canvas, color, cx, cy):
            """Draw the folder+plus icon centered at (cx, cy)."""
            s  = _NP_ICON / 24.0
            ox = cx - _NP_ICON / 2
            oy = cy - _NP_ICON / 2
            fp = [
                4*s+ox,     20*s+oy,
                2*s+ox,     20*s+oy,
                2*s+ox,     5*s+oy,
                2*s+ox,     3*s+oy,
                7.93*s+ox,  3*s+oy,
                9.6*s+ox,   3.9*s+oy,
                10.21*s+ox, 5*s+oy,
                12.1*s+ox,  6*s+oy,
                20*s+ox,    6*s+oy,
                22*s+ox,    8*s+oy,
                22*s+ox,    20*s+oy,
                20*s+ox,    20*s+oy,
                4*s+ox,     20*s+oy,
            ]
            canvas.create_line(*fp, fill=color, width=1.35, capstyle="round", joinstyle="round")
            canvas.create_line(12*s+ox, 10*s+oy, 12*s+ox, 16*s+oy,
                               fill=color, width=1.35, capstyle="round")
            canvas.create_line(9*s+ox, 13*s+oy, 15*s+ox, 13*s+oy,
                               fill=color, width=1.35, capstyle="round")

        # ── Bottom: single New Project button (always packed at bottom) ──
        self._sidebar_bottom_exp = tk.Frame(sidebar, bg="#fafbfc")
        self._sidebar_bottom_exp.pack(side="bottom", fill="x", padx=16, pady=(20, 32))

        _np_t    = [0.0]
        _np_anim = [None]

        btn_c = tk.Canvas(self._sidebar_bottom_exp, height=_NP_H+2, bg="#fafbfc",
                          highlightthickness=0, cursor=CURSOR_HAND)
        btn_c.pack(fill="x")
        self._np_btn_c    = btn_c
        self._np_btn_draw = None  # set after _np_draw is defined

        def _np_draw(color):
            btn_c.delete("all")
            W = btn_c.winfo_width()
            if W < 2:
                return
            H = _NP_H
            r = _NP_RADIUS; d = r * 2
            if getattr(self, "_sidebar_expanded", True):
                # Expanded: full-width pill button
                x1, y1, x2, y2 = 1, 1, W-1, H-1
                btn_c.create_arc(x1,    y1,    x1+d, y1+d, start=90,  extent=90,  fill=color, outline=color)
                btn_c.create_arc(x2-d,  y1,    x2,   y1+d, start=0,   extent=90,  fill=color, outline=color)
                btn_c.create_arc(x1,    y2-d,  x1+d, y2,   start=180, extent=90,  fill=color, outline=color)
                btn_c.create_arc(x2-d,  y2-d,  x2,   y2,   start=270, extent=90,  fill=color, outline=color)
                btn_c.create_rectangle(x1+r, y1,   x2-r, y2,   fill=color, outline=color)
                btn_c.create_rectangle(x1,   y1+r, x2,   y2-r, fill=color, outline=color)
                ix = (W - _np_block) / 2
                iy = H / 2
                _draw_np_icon(btn_c, _NP_FG, ix + _NP_ICON / 2, iy)
                btn_c.create_text(ix + _NP_ICON + _NP_GAP, iy, text=_NP_LABEL,
                                  fill=_NP_FG, anchor="w", font=_np_font)
            else:
                # Collapsed: square button centered in canvas
                sq = H
                sx = (W - sq) / 2
                x1, y1, x2, y2 = sx+1, 1, sx+sq-1, H-1
                btn_c.create_arc(x1,    y1,    x1+d, y1+d, start=90,  extent=90,  fill=color, outline=color)
                btn_c.create_arc(x2-d,  y1,    x2,   y1+d, start=0,   extent=90,  fill=color, outline=color)
                btn_c.create_arc(x1,    y2-d,  x1+d, y2,   start=180, extent=90,  fill=color, outline=color)
                btn_c.create_arc(x2-d,  y2-d,  x2,   y2,   start=270, extent=90,  fill=color, outline=color)
                btn_c.create_rectangle(x1+r, y1,   x2-r, y2,   fill=color, outline=color)
                btn_c.create_rectangle(x1,   y1+r, x2,   y2-r, fill=color, outline=color)
                _draw_np_icon(btn_c, _NP_FG, sx + sq / 2, H / 2)

        self._np_btn_draw = _np_draw

        def _np_animate(target):
            if _np_anim[0]:
                btn_c.after_cancel(_np_anim[0])
                _np_anim[0] = None
            def tick():
                diff = target - _np_t[0]
                if abs(diff) < 0.02:
                    _np_t[0] = target
                    _np_draw(_hex_interp(_NP_BG, _NP_HOVER, target))
                    _np_anim[0] = None
                    return
                _np_t[0] += diff * 0.3
                _np_draw(_hex_interp(_NP_BG, _NP_HOVER, _np_t[0]))
                _np_anim[0] = btn_c.after(16, tick)
            tick()

        def _np_poll():
            try:
                mx = btn_c.winfo_pointerx()
                my = btn_c.winfo_pointery()
                bx = btn_c.winfo_rootx()
                by = btn_c.winfo_rooty()
                bw = btn_c.winfo_width()
                bh = btn_c.winfo_height()
                over = bx <= mx <= bx + bw and by <= my <= by + bh
                target = 1.0 if over else 0.0
                if abs(_np_t[0] - target) > 0.01 and _np_anim[0] is None:
                    _np_animate(target)
            except tk.TclError:
                return
            btn_c.after(30, _np_poll)

        btn_c.bind("<Configure>", lambda e: _np_draw(_hex_interp(_NP_BG, _NP_HOVER, _np_t[0])))
        btn_c.bind("<Button-1>",  lambda e: self._new_project())
        btn_c.after(100, _np_poll)

        self._refresh_sidebar_projects()

    def _toggle_sidebar(self):
        self._sidebar_expanded = not self._sidebar_expanded
        if self._sidebar_expanded:
            self._sidebar_hdr_col.pack_forget()
            self._sidebar_hdr_exp.pack(fill="x", before=self._sidebar_exp_content)
            self._sidebar_proj_lbl.config(fg="#a3a3a3")
            self._sidebar_bottom_exp.pack_configure(padx=16, pady=(20, 32))
            self._sidebar.config(width=self._sidebar_exp_w)
        else:
            self._sidebar_hdr_exp.pack_forget()
            self._sidebar_hdr_col.pack(fill="x", before=self._sidebar_exp_content)
            self._sidebar_proj_lbl.config(fg="#fafbfc")   # invisible text, keeps spacing
            self._sidebar_bottom_exp.pack_configure(padx=12, pady=(20, 32))
            self._sidebar.config(width=self._sidebar_col_w)
        # Redraw NP button for new state
        if self._np_btn_draw:
            self._np_btn_c.event_generate("<Configure>")
        self._refresh_sidebar_projects()

    def _build_topbar(self):
        topbar = tk.Frame(self._content, bg="#ffffff", height=70)
        topbar.pack(fill="x")
        topbar.pack_propagate(False)
        self._topbar = topbar

        # ── Left: project name (4/12) ──────────────────────────────────
        left_wrap = tk.Frame(topbar, bg="#ffffff")
        left_wrap.place(relx=0, rely=0, relwidth=4/12, relheight=1)
        left_inner = tk.Frame(left_wrap, bg="#ffffff")
        left_inner.place(x=20, y=35, anchor="w")

        # ── Right: buttons (8/12) — flex-wrap ─────────────────────────
        right_wrap = tk.Frame(topbar, bg="#ffffff")
        right_wrap.place(relx=4/12, rely=0, relwidth=8/12, relheight=1)
        btn_frame = _FlowFrame(right_wrap, bg="#ffffff", vgap=6)
        btn_frame.place(x=0, y=18, relwidth=1.0, width=-12)

        def _sync_topbar_h(e, _tb=topbar):
            new_h = max(70, e.height + 36)   # 18px top + 18px bottom padding
            if _tb.winfo_height() != new_h:
                _tb.configure(height=new_h)
        btn_frame.bind("<Configure>", _sync_topbar_h, add="+")

        self._topbar_name_lbl = tk.Label(
            left_inner, text="",
            font=(FONT_FAMILY, 12, "bold"),
            bg="#ffffff", fg="#2c3b4e", anchor="w", justify="left",
        )
        self._topbar_name_lbl.pack(anchor="w")

        _sf = tk.Frame(left_inner, bg="#ffffff")
        _sf.pack(anchor="w")
        _f10 = (FONT_FAMILY, 9)
        self._stats_dm_lbl       = tk.Label(_sf, text="", font=_f10, bg="#ffffff", fg="#7799b9")
        self._stats_dot_lbl      = tk.Label(_sf, text="", font=_f10, bg="#ffffff", fg="#808080")
        self._stats_policies_lbl = tk.Label(_sf, text="", font=_f10, bg="#ffffff", fg="#a3a3a3")
        self._stats_pipe_lbl     = tk.Label(_sf, text="", font=_f10, bg="#ffffff", fg="#808080")
        self._stats_cells_lbl    = tk.Label(_sf, text="", font=_f10, bg="#ffffff", fg="#a3a3a3")
        for _lbl in (self._stats_dm_lbl, self._stats_dot_lbl, self._stats_policies_lbl,
                     self._stats_pipe_lbl, self._stats_cells_lbl):
            _lbl.pack(side="left")

        def _update_name_wrap(e):
            self._topbar_name_lbl.config(wraplength=max(100, e.width - 40))
        left_wrap.bind("<Configure>", _update_name_wrap)

        # ── Shared constants for all canvas buttons ───────────────────
        _EX_BG     = "#f5f7fa"
        _EX_HOVER  = "#eaeef4"
        _EX_BORDER = "#eaeef4"
        _EX_FG     = "#2c3b4e"
        _DEL_FG    = "#DC2626"
        _EX_H      = 33
        _EX_R      = 5
        _EX_ICON   = 14
        _EX_GAP    = 7
        _EX_PADX   = 20
        _ex_font   = tkFont.Font(family=FONT_FAMILY, size=11)

        def _make_icon_btn(label, cmd, icon_fn, fg):
            tw   = _ex_font.measure(label)
            w    = _EX_PADX + _EX_ICON + _EX_GAP + tw + _EX_PADX
            t    = [0.0]
            anim = [None]
            c = tk.Canvas(btn_frame, width=w+2, height=_EX_H+2,
                          bg="#ffffff", highlightthickness=0, cursor=CURSOR_HAND)
            btn_frame.add(c, padx=4)

            def draw(bg, _w=w, _tw=tw, _label=label, _c=c, _fg=fg):
                _c.delete("all")
                pts = _rrect_pts(1, 1, _w+1, _EX_H+1, _EX_R)
                _c.create_polygon(*pts, fill=bg, outline=_EX_BORDER, width=1)
                cx_ = 1 + _w // 2
                cy_ = 1 + _EX_H // 2
                ix  = cx_ - (_EX_ICON + _EX_GAP + _tw) // 2
                iy  = cy_ - _EX_ICON // 2
                icon_fn(_c, ix, iy, _fg)
                _c.create_text(ix + _EX_ICON + _EX_GAP, cy_,
                               text=_label, fill=_fg, anchor="w", font=_ex_font)

            def animate(target, _t=t, _anim=anim, _draw=draw, _c=c):
                if _anim[0]: _c.after_cancel(_anim[0]); _anim[0] = None
                def tick():
                    diff = target - _t[0]
                    if abs(diff) < 0.02:
                        _t[0] = target; _draw(_hex_interp(_EX_BG, _EX_HOVER, target))
                        _anim[0] = None; return
                    _t[0] += diff * 0.3
                    _draw(_hex_interp(_EX_BG, _EX_HOVER, _t[0]))
                    _anim[0] = _c.after(16, tick)
                tick()

            def poll(_t=t, _anim=anim, _animate=animate, _c=c, _w=w):
                try:
                    mx = _c.winfo_pointerx(); my = _c.winfo_pointery()
                    bx = _c.winfo_rootx();    by = _c.winfo_rooty()
                    over = bx <= mx <= bx + _w and by <= my <= by + _EX_H
                    tgt  = 1.0 if over else 0.0
                    if abs(_t[0] - tgt) > 0.01 and _anim[0] is None:
                        _animate(tgt)
                except tk.TclError:
                    return
                _c.after(30, poll)

            draw(_EX_BG)
            c.bind("<Button-1>", lambda e, fn=cmd: fn())
            c.after(100, poll)

        def _draw_pencil_icon(canvas, ox, oy, fg):
            s  = _EX_ICON / 24.0
            lw = 1.35
            kw = dict(fill=fg, width=lw, capstyle="round", joinstyle="round")
            body = [
                21*s+ox,  7*s+oy,  17*s+ox,  3*s+oy,
                 4*s+ox, 16*s+oy,   3*s+ox, 17*s+oy,
                 2*s+ox, 21*s+oy,   3*s+ox, 22*s+oy,
                 7*s+ox, 21*s+oy,   8*s+ox, 20*s+oy,
                21*s+ox,  7*s+oy,
            ]
            canvas.create_line(*body, **kw)
            canvas.create_line(15*s+ox, 5*s+oy, 19*s+ox, 9*s+oy, **kw)

        def _draw_trash_icon(canvas, ox, oy, fg):
            s  = _EX_ICON / 24.0
            lw = 1.35
            kw = dict(fill=fg, width=lw, capstyle="round", joinstyle="round")
            canvas.create_line(3*s+ox, 6*s+oy, 21*s+ox, 6*s+oy, **kw)
            canvas.create_line(
                19*s+ox,  6*s+oy, 19*s+ox, 20*s+oy,
                17*s+ox, 22*s+oy,  7*s+ox, 22*s+oy,
                 5*s+ox, 20*s+oy,  5*s+ox,  6*s+oy, **kw)
            canvas.create_line(
                 8*s+ox,  6*s+oy,  8*s+ox,  4*s+oy,
                10*s+ox,  2*s+oy, 14*s+ox,  2*s+oy,
                16*s+ox,  4*s+oy, 16*s+ox,  6*s+oy, **kw)

        def _draw_file_icon(canvas, ox, oy, fg, arrow_up=True):
            s  = _EX_ICON / 24.0
            lw = 1.35
            kw = dict(fill=fg, width=lw, capstyle="round", joinstyle="round")
            doc = [
                 6*s+ox,      22*s+oy,  4*s+ox,       20*s+oy,
                 4*s+ox,       4*s+oy,  6*s+ox,        2*s+oy,
                14*s+ox,       2*s+oy, 15.704*s+ox,  2.706*s+oy,
                19.292*s+ox, 6.294*s+oy, 20*s+ox,     8*s+oy,
                20*s+ox,      20*s+oy, 18*s+ox,       22*s+oy,
                 6*s+ox,      22*s+oy,
            ]
            canvas.create_line(*doc, **kw)
            canvas.create_line(14*s+ox, 2*s+oy, 14*s+ox, 7*s+oy, 20*s+ox, 7*s+oy, **kw)
            if arrow_up:
                canvas.create_line(12*s+ox, 18*s+oy, 12*s+ox, 12*s+oy, **kw)
                canvas.create_line(9*s+ox, 15*s+oy, 12*s+ox, 12*s+oy, 15*s+ox, 15*s+oy, **kw)
            else:
                canvas.create_line(12*s+ox, 12*s+oy, 12*s+ox, 18*s+oy, **kw)
                canvas.create_line(9*s+ox, 15*s+oy, 12*s+ox, 18*s+oy, 15*s+ox, 15*s+oy, **kw)

        # Rename Tab | Delete Tab
        _make_icon_btn("Rename Tab", self._rename_current_tab,
                       lambda c, ox, oy, fg: _draw_pencil_icon(c, ox, oy, fg), _EX_FG)
        _make_icon_btn("Delete Tab", self._remove_current_tab,
                       lambda c, ox, oy, fg: _draw_trash_icon(c, ox, oy, fg), _DEL_FG)

        btn_frame.add(tk.Frame(btn_frame, bg="#e6e6e6", width=1, height=_EX_H - 16), padx=6, is_sep=True)

        # Import Excel | Export Excel
        _make_icon_btn("Import Excel", self._import_excel,
                       lambda c, ox, oy, fg: _draw_file_icon(c, ox, oy, fg, arrow_up=False), _EX_FG)
        _make_icon_btn("Export Excel", self._export_excel,
                       lambda c, ox, oy, fg: _draw_file_icon(c, ox, oy, fg, arrow_up=True), _EX_FG)

        btn_frame.add(tk.Frame(btn_frame, bg="#e6e6e6", width=1, height=_EX_H - 16), padx=6, is_sep=True)

        # ── Run Analysis: canvas rounded button at far right ───────────
        _RA_BG    = "#2c3b4e"
        _RA_HOVER = "#37506d"
        _RA_FG    = "#f5f7fa"
        _RA_H     = 33
        _RA_R     = 5
        _RA_ICON  = 11
        _RA_GAP   = 7
        _RA_PADX  = 18

        _ra_font = tkFont.Font(family=FONT_FAMILY, size=11)
        _ra_tw   = _ra_font.measure("Run Analysis")
        _ra_w    = _RA_PADX + _RA_ICON + _RA_GAP + _ra_tw + _RA_PADX

        def _draw_play(canvas, ox, oy):
            s   = _RA_ICON / 24.0
            pts = [
                5*s+ox,      5*s+oy,
                8.008*s+ox,  3.272*s+oy,
                20.005*s+ox, 10.27*s+oy,
                20.008*s+ox, 13.728*s+oy,
                8.008*s+ox,  20.728*s+oy,
                5*s+ox,      19*s+oy,
            ]
            canvas.create_polygon(*pts, fill="", outline=_RA_FG,
                                  width=1.35, smooth=False)

        _ra_t    = [0.0]
        _ra_anim = [None]

        ra_c = tk.Canvas(btn_frame, width=_ra_w + 2, height=_RA_H + 2,
                         bg="#ffffff", highlightthickness=0, cursor=CURSOR_HAND)
        btn_frame.add(ra_c, padx=4)

        def _ra_draw(color):
            ra_c.delete("all")
            pts = _rrect_pts(1, 1, _ra_w + 1, _RA_H + 1, _RA_R)
            ra_c.create_polygon(*pts, fill=color, outline=color, width=0)
            cx = 1 + _ra_w // 2
            cy = 1 + _RA_H // 2
            content_w = _RA_ICON + _RA_GAP + _ra_tw
            ix = cx - content_w // 2
            iy = cy - _RA_ICON // 2
            _draw_play(ra_c, ix, iy)
            ra_c.create_text(
                ix + _RA_ICON + _RA_GAP, cy,
                text="Run Analysis", fill=_RA_FG, anchor="w", font=_ra_font,
            )

        def _ra_animate(target):
            if _ra_anim[0]:
                ra_c.after_cancel(_ra_anim[0])
                _ra_anim[0] = None
            def tick():
                diff = target - _ra_t[0]
                if abs(diff) < 0.02:
                    _ra_t[0] = target
                    _ra_draw(_hex_interp(_RA_BG, _RA_HOVER, target))
                    _ra_anim[0] = None
                    return
                _ra_t[0] += diff * 0.3
                _ra_draw(_hex_interp(_RA_BG, _RA_HOVER, _ra_t[0]))
                _ra_anim[0] = ra_c.after(16, tick)
            tick()

        def _ra_poll():
            try:
                mx = ra_c.winfo_pointerx(); my = ra_c.winfo_pointery()
                bx = ra_c.winfo_rootx();   by = ra_c.winfo_rooty()
                over = bx <= mx <= bx + _ra_w and by <= my <= by + _RA_H
                target = 1.0 if over else 0.0
                if abs(_ra_t[0] - target) > 0.01 and _ra_anim[0] is None:
                    _ra_animate(target)
            except tk.TclError:
                return
            ra_c.after(30, _ra_poll)

        _ra_draw(_RA_BG)
        ra_c.bind("<Button-1>", lambda e: self._run_aggregation())
        ra_c.after(100, _ra_poll)

        # ── Bottom divider ─────────────────────────────────────────────
        tk.Frame(self._content, bg="#e6e6e6", height=1).pack(fill="x")

    def _update_topbar(self):
        """Refresh the top bar project name and stats for the current project."""
        def _clear_stats():
            for lbl in (self._stats_dm_lbl, self._stats_dot_lbl,
                        self._stats_policies_lbl, self._stats_pipe_lbl,
                        self._stats_cells_lbl):
                lbl.config(text="")

        proj = self._current_project()
        if proj is None:
            self._topbar_name_lbl.config(text="")
            _clear_stats()
            return

        self._topbar_name_lbl.config(text=proj.name)

        if not proj.matrices:
            self._stats_dm_lbl.config(text="No decision-makers yet")
            for lbl in (self._stats_dot_lbl, self._stats_policies_lbl,
                        self._stats_pipe_lbl, self._stats_cells_lbl):
                lbl.config(text="")
            return

        if getattr(proj, "_in_analysis_view", False):
            n_dms      = len(proj.matrices)
            n_policies = len(proj.matrices[0].policies)
            pol_text   = f"{n_policies} {'policy' if n_policies == 1 else 'policies'}"
            dm_text    = f"{n_dms} decision-maker{'s' if n_dms != 1 else ''}"
            self._stats_dm_lbl.config(text=dm_text)
            self._stats_dot_lbl.config(text="  .  ")
            self._stats_policies_lbl.config(text=pol_text)
            self._stats_pipe_lbl.config(text="  |  ")
            self._stats_cells_lbl.config(text="Analysis")
            return

        # Get currently selected DM tab
        selected_tab = proj.notebook.select()
        matrix = next((m for m in proj.matrices if str(m._tab) == selected_tab),
                      proj.matrices[0])

        n_policies = len(matrix.policies)
        n_cells    = matrix.total_cells()
        pol_text   = f"{n_policies} {'policy' if n_policies == 1 else 'policies'}"

        self._stats_dm_lbl.config(text=f"{matrix.decision_maker}'s workspace")
        self._stats_dot_lbl.config(text="  .  ")
        self._stats_policies_lbl.config(text=pol_text)
        self._stats_pipe_lbl.config(text="  |  ")
        self._stats_cells_lbl.config(text=f"{n_cells} cells to fill")

    def _on_project_tab_changed(self, _event=None):
        proj = self._current_project()
        if proj is None:
            self._update_topbar()
            return
        self._update_topbar()
        if getattr(proj, "_in_analysis_view", False):
            self._refresh_an_underlines(proj)
            return
        self._refresh_dm_underlines(proj)
        self._ensure_current_dm_widget(proj)

    def _select_project(self, proj: Project):
        if proj is None:
            return
        self._proj_nb.select(proj.frame)
        self.root.update_idletasks()
        self._update_topbar()
        if getattr(proj, "_in_analysis_view", False):
            self._refresh_an_underlines(proj)
            return
        self._refresh_dm_underlines(proj)
        self._ensure_current_dm_widget(proj)
        proj.frame.update_idletasks()

    def _build_project_notebook(self):
        self._proj_nb_container = tk.Frame(self._content, bg=COLOR_BG)
        pack_kwargs = dict(fill="both", expand=True)
        if hasattr(self, "_statusbar"):
            pack_kwargs["before"] = self._statusbar
        self._proj_nb_container.pack(**pack_kwargs)
        self._proj_nb = ttk.Notebook(self._proj_nb_container, style="Headless.TNotebook")
        self._proj_nb.pack(fill="both", expand=True)

        self._empty_panel_w = 500
        self._empty_resize_job = None
        self._empty_label = self._build_empty_state(self._proj_nb_container, 500)
        self._empty_label.place(x=0, y=0, relwidth=1, relheight=1)
        self._proj_nb_container.bind("<Configure>", self._on_empty_resize)

    def _on_empty_resize(self, event):
        if self._empty_resize_job:
            self.root.after_cancel(self._empty_resize_job)
        self._empty_resize_job = self.root.after(
            120, lambda w=event.width: self._rebuild_empty_state(w)
        )

    def _rebuild_empty_state(self, avail_w):
        self._empty_resize_job = None
        new_w = max(280, min(500, avail_w - 80))
        if abs(new_w - self._empty_panel_w) < 10:
            return
        self._empty_panel_w = new_w
        visible = self._empty_label.winfo_ismapped()
        self._empty_label.destroy()
        self._empty_label = self._build_empty_state(self._proj_nb_container, new_w)
        if visible:
            self._empty_label.place(x=0, y=0, relwidth=1, relheight=1)

    def _build_empty_state(self, parent, panel_w=500):
        """Scrollable centered instruction panel shown when no project exists."""
        PANEL_W  = panel_w
        BG       = COLOR_BG

        # Full-size wrapper — fills the container via place(relwidth=1, relheight=1)
        wrapper = tk.Frame(parent, bg=BG)
        _sc = tk.Canvas(wrapper, bg=BG, highlightthickness=0)
        _sc.pack(fill="both", expand=True)

        _mode    = [None]   # "normal" | "compact"
        _win_ref = [None]
        _outer_ref = [None]

        def _build_inner(compact):
            if _outer_ref[0] is not None:
                if _win_ref[0] is not None:
                    _sc.delete(_win_ref[0])
                _outer_ref[0].destroy()

            outer = tk.Frame(_sc, bg=BG)
            _outer_ref[0] = outer

            # ── Responsive size tokens ───────────────────────────────────
            if compact:
                BADGE_SIZE_  = 44
                ICON_SIZE_   = 20
                BADGE_R_     = 7
                LW_          = 1.5
                BADGE_PADY   = 8
                TITLE_SIZE   = 17
                TITLE_PADY   = (0, 6)
                SUB_PADY     = (0, 10)
                WF_PY_       = 14
                LBL_GAP_     = 6
                STEP_H_      = 40
                STEP_GAP_    = 5
                WF_PADY      = (0, 10)
            else:
                BADGE_SIZE_  = 58
                ICON_SIZE_   = 26
                BADGE_R_     = 8
                LW_          = 1.6
                BADGE_PADY   = 18
                TITLE_SIZE   = 23
                TITLE_PADY   = (0, 10)
                SUB_PADY     = (0, 22)
                WF_PY_       = 26
                LBL_GAP_     = 12
                STEP_H_      = 55
                STEP_GAP_    = 8
                WF_PADY      = (0, 22)

        # ── Icon badge ──────────────────────────────────────────────────
            BADGE_SIZE = BADGE_SIZE_
            ICON_SIZE  = ICON_SIZE_
            BADGE_BG   = "#eaeef4"
            ICON_COLOR = "#30455c"
            BADGE_R    = BADGE_R_
            LW         = LW_

            badge_c = tk.Canvas(outer, width=BADGE_SIZE+2, height=BADGE_SIZE+2,
                                bg=BG, highlightthickness=0)
            badge_c.pack(pady=(0, BADGE_PADY))

            # Rounded-rect badge background (inset 1px so corners aren't clipped)
            pts = _rrect_pts(1, 1, BADGE_SIZE+1, BADGE_SIZE+1, BADGE_R)
            badge_c.create_polygon(*pts, fill=BADGE_BG, outline=BADGE_BG)

            # Network icon — Lucide "network", 24×24 viewBox scaled to ICON_SIZE
            s  = ICON_SIZE / 24.0
            ox = 1 + (BADGE_SIZE - ICON_SIZE) / 2
            oy = 1 + (BADGE_SIZE - ICON_SIZE) / 2
            lkw = dict(fill=ICON_COLOR, width=LW, capstyle="round", joinstyle="round")

            # rect x="16" y="16" width="6" height="6" rx="1"
            pts = _rrect_pts(16*s+ox, 16*s+oy, 22*s+ox, 22*s+oy, 1*s)
            badge_c.create_polygon(*pts, fill="", outline=ICON_COLOR, width=LW)
            # rect x="2" y="16" width="6" height="6" rx="1"
            pts = _rrect_pts(2*s+ox, 16*s+oy, 8*s+ox, 22*s+oy, 1*s)
            badge_c.create_polygon(*pts, fill="", outline=ICON_COLOR, width=LW)
            # rect x="9" y="2" width="6" height="6" rx="1"
            pts = _rrect_pts(9*s+ox, 2*s+oy, 15*s+ox, 8*s+oy, 1*s)
            badge_c.create_polygon(*pts, fill="", outline=ICON_COLOR, width=LW)
            # path M5 16 v-3 a1 1 0 0 1 1-1 h12 a1 1 0 0 1 1 1 v3
            badge_c.create_line(
                5*s+ox, 16*s+oy, 5*s+ox, 13*s+oy,
                6*s+ox, 12*s+oy, 18*s+ox, 12*s+oy,
                19*s+ox, 13*s+oy, 19*s+ox, 16*s+oy,
                **lkw
            )
            # path M12 12 V8
            badge_c.create_line(12*s+ox, 12*s+oy, 12*s+ox, 8*s+oy, **lkw)

            # ── Title ────────────────────────────────────────────────────────
            tk.Label(
                outer,
                text="Start a new policy coherence assessment",
                font=(FONT_FAMILY, TITLE_SIZE),
                bg=BG, fg="#1f2937",
                wraplength=PANEL_W, justify="center",
            ).pack(pady=TITLE_PADY)

            # ── Subtitle ─────────────────────────────────────────────────────
            tk.Label(
                outer,
                text=(
                    "Create a project to define policies, add decision-makers, and evaluate "
                    "interactions across expert perspectives. The tool supports structured matrix "
                    "input, aggregation methods, and advanced analytical outputs."
                ),
                font=(FONT_FAMILY, 12),
                bg=BG, fg="#808080",
                wraplength=PANEL_W, justify="left",
            ).pack(pady=SUB_PADY, anchor="w")

        # ── Workflow box (rounded border + drop shadow via Canvas) ────────
            WF_R        = 6
            WF_PX       = 28    # inner horizontal padding
            WF_PY       = WF_PY_
            WF_W        = PANEL_W
            WF_BG       = "#fafbfc"
            WF_BORDER   = "#f5f7fa"
            WF_INNER_W  = WF_W - 2 * WF_PX - 2
            SHADOW_OFF  = 1.25
            SHADOW_CLR  = "#e8eaed"

            STEP_H      = STEP_H_
            STEP_R      = 4
            STEP_BG     = "#ffffff"
            STEP_BORDER = "#f5f5f5"
            STEP_GAP    = STEP_GAP_
            STEPS = (
            "Create a project",
            "Add decision-makers",
            "Define policies",
            "Complete the interaction matrix",
            "Run aggregation and analysis",
        )
            N_STEPS = len(STEPS)
            LBL_H   = 22
            LBL_GAP = LBL_GAP_
            WF_H    = (WF_PY + LBL_H + LBL_GAP
                       + N_STEPS * STEP_H + (N_STEPS - 1) * STEP_GAP
                       + WF_PY)

            wf_c = tk.Canvas(outer, width=WF_W + SHADOW_OFF, height=WF_H + SHADOW_OFF,
                             bg=BG, highlightthickness=0)
            wf_c.pack(pady=WF_PADY)

        # Drop shadow (drawn first, offset by SHADOW_OFF)
            s_pts = _rrect_pts(1 + SHADOW_OFF, 1 + SHADOW_OFF,
                               WF_W - 1 + SHADOW_OFF, WF_H - 1 + SHADOW_OFF, WF_R)
            wf_c.create_polygon(*s_pts, fill=SHADOW_CLR, outline="")

            # Main box
            pts = _rrect_pts(1, 1, WF_W - 1, WF_H - 1, WF_R)
            wf_c.create_polygon(*pts, fill=WF_BG, outline=WF_BORDER, width=1)

            wf_inner = tk.Frame(wf_c, bg=WF_BG, width=WF_INNER_W)
            wf_c.create_window(WF_PX + 1, WF_PY + 1, window=wf_inner, anchor="nw")

            tk.Label(
                wf_inner, text="WORKFLOW",
                font=(FONT_FAMILY, 12),
                bg=WF_BG, fg="#a3a3a3",
                anchor="w",
            ).pack(fill="x", pady=(0, LBL_GAP))

            # ── Step icon constants ───────────────────────────────────────────
            ICON_COLOR   = "#2c3b4e"
            ICON_LW      = 1.35
            BADGE_SZ     = 28
            BADGE_R_ICN  = 5
            BADGE_BG_ICN = "#f5f7fa"
            ICON_SZ      = 14
            SC           = ICON_SZ / 24.0          # scale factor from 24px viewBox
            BADGE_X      = 40                       # badge left edge in step canvas
            BADGE_Y      = (STEP_H - BADGE_SZ) / 2
            IOX          = BADGE_X + (BADGE_SZ - ICON_SZ) / 2
            IOY          = BADGE_Y + (BADGE_SZ - ICON_SZ) / 2
            TEXT_X       = BADGE_X + BADGE_SZ + 12

            def _ilkw():
                return dict(fill=ICON_COLOR, width=ICON_LW, capstyle="round", joinstyle="round")

            def _icon_folder_plus(c, ox, oy):
                lkw = _ilkw()
                fp = [(2,5),(2,18),(4,20),(20,20),(22,18),(22,8),(20,6),
                      (12.1,6),(10.41,5.1),(9.6,3.9),(7.93,3),(4,3),(2,5)]
                flat = [v for x, y in fp for v in (x*SC+ox, y*SC+oy)]
                c.create_line(*flat, **lkw)
                c.create_line(12*SC+ox, 10*SC+oy, 12*SC+ox, 16*SC+oy, **lkw)
                c.create_line( 9*SC+ox, 13*SC+oy, 15*SC+ox, 13*SC+oy, **lkw)

            def _icon_user_plus(c, ox, oy):
                lkw = _ilkw()
                body = [(16,21),(16,19),(15,17),(12,16),(6,16),(3,17),(2,19),(2,21)]
                flat = [v for x, y in body for v in (x*SC+ox, y*SC+oy)]
                c.create_line(*flat, **lkw)
                r = 4 * SC
                c.create_oval(9*SC+ox-r, 7*SC+oy-r, 9*SC+ox+r, 7*SC+oy+r,
                              outline=ICON_COLOR, width=ICON_LW, fill="")
                c.create_line(19*SC+ox,  8*SC+oy, 19*SC+ox, 14*SC+oy, **lkw)
                c.create_line(22*SC+ox, 11*SC+oy, 16*SC+ox, 11*SC+oy, **lkw)

            def _icon_file_text(c, ox, oy):
                lkw = _ilkw()
                fp = [(6,22),(4,20),(4,4),(6,2),(14,2),(15.7,2.7),
                      (19.6,6.3),(20,8),(20,20),(18,22),(6,22)]
                flat = [v for x, y in fp for v in (x*SC+ox, y*SC+oy)]
                c.create_line(*flat, **lkw)
                c.create_line(14*SC+ox, 2*SC+oy, 14*SC+ox, 7*SC+oy, 20*SC+ox, 7*SC+oy, **lkw)
                c.create_line(10*SC+ox,  9*SC+oy,  8*SC+ox,  9*SC+oy, **lkw)
                c.create_line(16*SC+ox, 13*SC+oy,  8*SC+ox, 13*SC+oy, **lkw)
                c.create_line(16*SC+ox, 17*SC+oy,  8*SC+ox, 17*SC+oy, **lkw)

            def _icon_square_pointer(c, ox, oy):
                lkw = _ilkw()
                sq = [(21,11),(21,5),(19,3),(5,3),(3,5),(3,19),(5,21),(11,21)]
                flat = [v for x, y in sq for v in (x*SC+ox, y*SC+oy)]
                c.create_line(*flat, **lkw)
                cur = [(12.034,12.681),(21.681,15.534),(18.204,17.545),(16.477,21.648),(12.034,12.681)]
                flat2 = [v for x, y in cur for v in (x*SC+ox, y*SC+oy)]
                c.create_line(*flat2, **lkw)

            def _icon_chart_network(c, ox, oy):
                lkw = _ilkw()
                c.create_line(3*SC+ox, 3*SC+oy, 3*SC+ox, 19*SC+oy,
                              5*SC+ox, 21*SC+oy, 21*SC+ox, 21*SC+oy, **lkw)
                c.create_line(13.11*SC+ox, 7.664*SC+oy, 14.89*SC+ox, 10.336*SC+oy, **lkw)
                c.create_line(14.162*SC+ox, 12.788*SC+oy, 10.838*SC+ox, 14.212*SC+oy, **lkw)
                c.create_line(20*SC+ox, 4*SC+oy, 13.94*SC+ox, 5.515*SC+oy, **lkw)
                r = 2 * SC
                for nx, ny in ((12, 6), (16, 12), (9, 15)):
                    c.create_oval(nx*SC+ox-r, ny*SC+oy-r, nx*SC+ox+r, ny*SC+oy+r,
                                  outline=ICON_COLOR, width=ICON_LW, fill="")

            ICON_FUNCS = [
                _icon_folder_plus, _icon_user_plus, _icon_file_text,
                _icon_square_pointer, _icon_chart_network,
            ]

            for i, step in enumerate(STEPS):
                pady = (0, STEP_GAP) if i < N_STEPS - 1 else 0

                step_c = tk.Canvas(wf_inner, width=WF_INNER_W, height=STEP_H,
                                   bg=WF_BG, highlightthickness=0)
                step_c.pack(fill="x", pady=pady)

                sp = _rrect_pts(1, 1, WF_INNER_W - 1, STEP_H - 1, STEP_R)
                step_c.create_polygon(*sp, fill=STEP_BG, outline=STEP_BORDER, width=1)

                # Step number
                step_c.create_text(
                    14, STEP_H / 2,
                    text=f"{i + 1:02d}", anchor="w",
                    font=(FONT_FAMILY, 10),
                    fill="#d3d3d3",
                )

                # Icon badge + icon
                bp = _rrect_pts(BADGE_X, BADGE_Y, BADGE_X + BADGE_SZ, BADGE_Y + BADGE_SZ, BADGE_R_ICN)
                step_c.create_polygon(*bp, fill=BADGE_BG_ICN, outline="")
                ICON_FUNCS[i](step_c, IOX, IOY)

                # Step text
                step_c.create_text(
                    TEXT_X, STEP_H / 2,
                    text=step, anchor="w",
                    font=(FONT_FAMILY, 13),
                    fill="#1f2937",
                )

            # ── New Project button ────────────────────────────────────────────
            BTN_BG    = "#426387"
            BTN_HOVER = "#30455c"
            BTN_FG    = "#eaeef4"
            BTN_H     = 35
            BTN_R     = 4
            BTN_ICON  = 16
            BTN_GAP   = 7
            BTN_PADX  = 22
            BTN_LABEL = "New Project"

            btn_font = tkFont.Font(family=FONT_FAMILY, size=11)
            tw       = btn_font.measure(BTN_LABEL)
            btn_w    = BTN_PADX + BTN_ICON + BTN_GAP + tw + BTN_PADX

            t    = [0.0]
            anim = [None]

            btn_c = tk.Canvas(outer, width=btn_w+2, height=BTN_H+2,
                              bg=BG, highlightthickness=0, cursor=CURSOR_HAND)
            btn_c.pack()

            def draw_btn(bg):
                btn_c.delete("all")
                pts = _rrect_pts(1, 1, btn_w+1, BTN_H+1, BTN_R)
                btn_c.create_polygon(*pts, fill=bg, outline=bg)
                s2  = BTN_ICON / 24.0
                ix  = 1 + BTN_PADX
                iy  = 1 + (BTN_H - BTN_ICON) / 2
                cy  = 1 + BTN_H / 2
                pkw = dict(fill=BTN_FG, width=1.35, capstyle="round")
                btn_c.create_line(5*s2+ix, 12*s2+iy, 19*s2+ix, 12*s2+iy, **pkw)
                btn_c.create_line(12*s2+ix, 5*s2+iy, 12*s2+ix, 19*s2+iy, **pkw)
                btn_c.create_text(
                    ix + BTN_ICON + BTN_GAP, cy,
                    text=BTN_LABEL, fill=BTN_FG, anchor="w", font=btn_font,
                )

            def animate_btn(target):
                if anim[0]: btn_c.after_cancel(anim[0]); anim[0] = None
                def tick():
                    diff = target - t[0]
                    if abs(diff) < 0.02:
                        t[0] = target; draw_btn(_hex_interp(BTN_BG, BTN_HOVER, target))
                        anim[0] = None; return
                    t[0] += diff * 0.3
                    draw_btn(_hex_interp(BTN_BG, BTN_HOVER, t[0]))
                    anim[0] = btn_c.after(16, tick)
                tick()

            def poll_btn():
                try:
                    mx = btn_c.winfo_pointerx(); my = btn_c.winfo_pointery()
                    bx = btn_c.winfo_rootx();    by = btn_c.winfo_rooty()
                    over = bx <= mx <= bx + btn_w and by <= my <= by + BTN_H
                    tgt  = 1.0 if over else 0.0
                    if abs(t[0] - tgt) > 0.01 and anim[0] is None:
                        animate_btn(tgt)
                except tk.TclError:
                    return
                btn_c.after(30, poll_btn)

            draw_btn(BTN_BG)
            btn_c.bind("<Button-1>", lambda e: self._new_project())
            btn_c.after(100, poll_btn)

            outer.bind("<Configure>", lambda e: _sc.after(10, _layout))
            _win_ref[0] = _sc.create_window(0, 0, window=outer, anchor="nw")

        def _layout(ev=None):
            cw = _sc.winfo_width()
            ch = _sc.winfo_height()
            compact    = ch < 680
            new_mode   = "compact" if compact else "normal"
            if new_mode != _mode[0]:
                _mode[0] = new_mode
                _build_inner(compact)
                _sc.after(10, _layout)
                return
            outer = _outer_ref[0]
            if outer is None:
                return
            ow = outer.winfo_reqwidth()
            oh = outer.winfo_reqheight()
            x  = max(0, (cw - ow) // 2)
            y  = max(20, (ch - oh) // 2)
            _sc.coords(_win_ref[0], x, y)
            _sc.configure(scrollregion=(0, 0, cw, max(ch, y + oh + 20)))

        _sc.bind("<Configure>", lambda e: _sc.after(10, _layout))

        def _on_wheel(event):
            _sc.yview_scroll(int(-1 * (event.delta / 120)), "units")
        _sc.bind("<MouseWheel>", _on_wheel)

        return wrapper

    def _build_statusbar(self):
        self._status_var = tk.StringVar(value="Ready")
        bar = tk.Frame(self._content, bg="#ffffff", height=35)
        bar.pack(fill="x", side="bottom")
        bar.pack_propagate(False)
        tk.Frame(bar, bg="#e6e6e6", height=1).pack(fill="x", side="top")
        tk.Label(bar, textvariable=self._status_var,
                 font=tkFont.Font(family=FONT_FAMILY, size=10, weight="normal"),
                 bg="#ffffff", fg="#a6bcd3",
                 anchor="w", padx=12).pack(fill="both", expand=True)
        self._statusbar = bar

    # ==================================================================
    # Project management
    # ==================================================================

    def _new_project(self):
        existing = [p.name for p in self.projects]
        dlg = ProjectSetupDialog(
            self.root,
            title="Create Project",
            include_project_name=True,
            existing_project_names=existing,
        )
        self.root.wait_window(dlg)
        if dlg.result:
            data = dlg.result
            if data.get("action") == "import_excel":
                self._import_excel_new_project()
                return
            proj = self._add_project(data["project_name"])
            for dm_name in data["decision_makers"]:
                matrix = make_empty_matrix(dm_name, data["policies"])
                proj.matrices.append(matrix)
                self._create_dm_tab(proj, matrix)
            self._proj_nb.select(proj.frame)
            self._refresh_sidebar_projects()
            self._set_status(
                f'Project "{proj.name}" created with {len(data["decision_makers"])} decision-maker'
                f'{"s" if len(data["decision_makers"]) != 1 else ""}.'
            )
            self._update_topbar()

    def _add_project(self, name: str):
        """Create a new Project and its top-level notebook tab."""
        self._empty_label.place_forget()

        proj = Project(name=name)

        # Outer frame inside the project notebook
        outer = tk.Frame(self._proj_nb, bg=COLOR_BG)
        self._proj_nb.add(outer, text=f"  {name}  ")
        self._proj_nb.select(outer)
        proj.frame = outer

        # ── View switcher bar (persistent pill toggle, right-aligned) ────────
        SW_BG     = "#ffffff"
        SW_H      = 48                     # switcher bar fixed height
        switcher  = tk.Frame(outer, bg=SW_BG, height=SW_H)
        switcher.pack(fill="x")
        switcher.pack_propagate(False)
        tk.Frame(outer, bg="#e8eaed", height=1).pack(fill="x")

        TRACK_BG  = "#a6bcd3"
        THUMB_BG  = "#e8eef4"
        ACT_FG    = "#30455c"
        INACT_FG  = "#ffffff"
        PILL_H    = 30
        THUMB_PAD = 3
        ICON_SZ   = 16
        SEG_W     = ICON_SZ + 28          # icon + 14px padding each side
        PILL_W    = SEG_W * 2

        # Custom tab label strips — dm_tab_bar is horizontally scrollable
        # Reserve right space so tabs don't slide under the pill toggle
        tk.Frame(switcher, width=PILL_W + 24, bg=SW_BG).pack(side="right")

        _dm_scroll_c = tk.Canvas(switcher, bg=SW_BG, highlightthickness=0,
                                 xscrollincrement=40)
        _dm_scroll_c.pack(side="left", fill="both", expand=True)
        dm_tab_bar  = tk.Frame(_dm_scroll_c, bg=SW_BG)
        _dm_bar_win = _dm_scroll_c.create_window((0, 0), window=dm_tab_bar, anchor="nw")

        _dm_scroll_c.bind("<Configure>",
            lambda e: _dm_scroll_c.itemconfig(_dm_bar_win, height=e.height))
        dm_tab_bar.bind("<Configure>",
            lambda e: _dm_scroll_c.configure(scrollregion=(0, 0, e.width, e.height)))
        def _scroll_fn(e, c=_dm_scroll_c, d=None):
            raw = getattr(e, "delta", 0)
            if not raw:
                return
            steps = raw if sys.platform == "darwin" else raw / 120.0
            _animate_canvas_x_scroll(c, d, -steps * 140, duration=220)
        _dm_scroll_c.bind("<MouseWheel>", _scroll_fn)
        dm_tab_bar.bind("<MouseWheel>", _scroll_fn)

        # ── Drag-to-scroll with momentum on the tab bar ───────────────
        _dm_drag = {"x": 0, "dragging": False, "vx": 0.0,
                    "history": [], "anim": [None]}
        _scroll_fn = lambda e, c=_dm_scroll_c, d=_dm_drag: (
            _animate_canvas_x_scroll(
                c, d,
                -(getattr(e, "delta", 0) if sys.platform == "darwin"
                  else getattr(e, "delta", 0) / 120.0) * 140,
                duration=220,
            ) if getattr(e, "delta", 0) else None
        )
        _dm_scroll_c.bind("<MouseWheel>", _scroll_fn)
        dm_tab_bar.bind("<MouseWheel>", _scroll_fn)

        def _dm_drag_start(e, d=_dm_drag, c=_dm_scroll_c):
            if d["anim"][0]:
                c.after_cancel(d["anim"][0])
                d["anim"][0] = None
            d["x"]       = e.x_root
            d["vx"]      = 0.0
            d["history"] = [(e.time, e.x_root)]
            d["dragging"] = False

        def _dm_drag_motion(e, c=_dm_scroll_c, d=_dm_drag):
            dx = d["x"] - e.x_root
            d["x"] = e.x_root

            # Keep a 80 ms position history for stable velocity estimation
            d["history"].append((e.time, e.x_root))
            d["history"] = [(t, x) for t, x in d["history"]
                            if e.time - t <= 120]
            if len(d["history"]) >= 2:
                t0, x0 = d["history"][0]
                t1, x1 = d["history"][-1]
                dt = max(1, t1 - t0)
                d["vx"] = (x0 - x1) / dt     # px / ms, positive = scroll right

            if abs(dx) > 1:
                d["dragging"] = True
            if not d["dragging"]:
                return
            _scroll_canvas_x_by_pixels(c, dx)

        def _dm_drag_end(e, c=_dm_scroll_c, d=_dm_drag):
            vx = [d["vx"]]

            def _coast():
                if abs(vx[0]) < 0.03:
                    d["anim"][0] = None
                    d["dragging"] = False
                    return
                if not _scroll_canvas_x_by_pixels(c, vx[0] * 16):
                    d["anim"][0] = None
                    d["dragging"] = False
                    return
                vx[0] *= 0.95
                d["anim"][0] = c.after(16, _coast)

            if abs(d["vx"]) > 0.03:
                d["anim"][0] = c.after(16, _coast)
            else:
                d["dragging"] = False

        for _w in (_dm_scroll_c, dm_tab_bar):
            _w.bind("<Button-1>",        _dm_drag_start)
            _w.bind("<B1-Motion>",       _dm_drag_motion)
            _w.bind("<ButtonRelease-1>", _dm_drag_end)

        _an_scroll_c = tk.Canvas(switcher, bg=SW_BG, highlightthickness=0,
                                 xscrollincrement=40)
        an_tab_bar  = tk.Frame(_an_scroll_c, bg=SW_BG)
        _an_bar_win = _an_scroll_c.create_window((0, 0), window=an_tab_bar, anchor="nw")

        _an_scroll_c.bind("<Configure>",
            lambda e: _an_scroll_c.itemconfig(_an_bar_win, height=e.height))
        an_tab_bar.bind("<Configure>",
            lambda e: _an_scroll_c.configure(scrollregion=(0, 0, e.width, e.height)))
        _an_scroll_fn = lambda e, c=_an_scroll_c, d=None: (
            _animate_canvas_x_scroll(
                c, d,
                -(getattr(e, "delta", 0) if sys.platform == "darwin"
                  else getattr(e, "delta", 0) / 120.0) * 140,
                duration=220,
            ) if getattr(e, "delta", 0) else None
        )
        _an_scroll_c.bind("<MouseWheel>", _an_scroll_fn)
        an_tab_bar.bind("<MouseWheel>", _an_scroll_fn)

        _an_drag = {"x": 0, "dragging": False, "vx": 0.0,
                    "history": [], "anim": [None]}
        _an_scroll_fn = lambda e, c=_an_scroll_c, d=_an_drag: (
            _animate_canvas_x_scroll(
                c, d,
                -(getattr(e, "delta", 0) if sys.platform == "darwin"
                  else getattr(e, "delta", 0) / 120.0) * 140,
                duration=220,
            ) if getattr(e, "delta", 0) else None
        )
        _an_scroll_c.bind("<MouseWheel>", _an_scroll_fn)
        an_tab_bar.bind("<MouseWheel>", _an_scroll_fn)

        def _an_drag_start(e, d=_an_drag, c=_an_scroll_c):
            if d["anim"][0]:
                c.after_cancel(d["anim"][0])
                d["anim"][0] = None
            d["x"] = e.x_root
            d["vx"] = 0.0
            d["history"] = [(e.time, e.x_root)]
            d["dragging"] = False

        def _an_drag_motion(e, c=_an_scroll_c, d=_an_drag):
            dx = d["x"] - e.x_root
            d["x"] = e.x_root
            d["history"].append((e.time, e.x_root))
            d["history"] = [(t, x) for t, x in d["history"] if e.time - t <= 120]
            if len(d["history"]) >= 2:
                t0, x0 = d["history"][0]
                t1, x1 = d["history"][-1]
                dt = max(1, t1 - t0)
                d["vx"] = (x0 - x1) / dt
            if abs(dx) > 1:
                d["dragging"] = True
            if not d["dragging"]:
                return
            _scroll_canvas_x_by_pixels(c, dx)

        def _an_drag_end(e, c=_an_scroll_c, d=_an_drag):
            vx = [d["vx"]]

            def _coast():
                if abs(vx[0]) < 0.03:
                    d["anim"][0] = None
                    d["dragging"] = False
                    return
                if not _scroll_canvas_x_by_pixels(c, vx[0] * 16):
                    d["anim"][0] = None
                    d["dragging"] = False
                    return
                vx[0] *= 0.95
                d["anim"][0] = c.after(16, _coast)

            if abs(d["vx"]) > 0.03:
                d["anim"][0] = c.after(16, _coast)
            else:
                d["dragging"] = False

        for _w in (_an_scroll_c, an_tab_bar):
            _w.bind("<Button-1>", _an_drag_start)
            _w.bind("<B1-Motion>", _an_drag_motion)
            _w.bind("<ButtonRelease-1>", _an_drag_end)

        proj._dm_tab_bar     = dm_tab_bar
        proj._dm_scroll_c    = _dm_scroll_c
        proj._an_scroll_c    = _an_scroll_c
        proj._an_tab_bar     = an_tab_bar
        proj._dm_tab_entries = []   # [(nb_tab_widget, label, underline, container)]
        proj._an_tab_entries = []   # [(tab_id_str,   label, underline, container)]
        proj._dm_drag        = _dm_drag
        proj._dm_drag_start  = _dm_drag_start
        proj._dm_drag_motion = _dm_drag_motion
        proj._an_drag        = _an_drag
        proj._an_drag_start  = _an_drag_start
        proj._an_drag_motion = _an_drag_motion

        toggle_c = tk.Canvas(switcher, width=PILL_W, height=PILL_H,
                             bg=SW_BG, highlightthickness=0, bd=0, cursor=CURSOR_HAND)
        toggle_c.place(relx=1.0, rely=0.5, anchor="e", x=-12)

        def _fill_pill(c, x1, y1, x2, y2, fill):
            """Reliable pill fill: two ovals on the ends + rectangle in the middle."""
            r = (y2 - y1) / 2
            c.create_oval(x1, y1, x1 + 2*r, y2, fill=fill, outline="")
            c.create_oval(x2 - 2*r, y1, x2, y2, fill=fill, outline="")
            c.create_rectangle(x1 + r, y1, x2 - r, y2, fill=fill, outline="")

        _s      = ICON_SZ / 24.0
        ICON_LW = 1.4

        def _icon_circle_user(c, ox, oy, color):
            """Lucide 'user': head circle (cx=12,cy=7,r=4) + body path with proper arc corners.
            Path: M19 21 v-2 a4 4 0 0 0 -4 -4 H9 a4 4 0 0 0 -4 4 v2"""
            lkw = dict(fill=color, width=ICON_LW, capstyle="round", joinstyle="round")
            # Head: circle cx=12, cy=7, r=4
            cr = 4 * _s
            c.create_oval(12*_s+ox-cr, 7*_s+oy-cr, 12*_s+ox+cr, 7*_s+oy+cr,
                          outline=color, fill="", width=ICON_LW)
            # Body: build polyline from straight segments + sampled arc curves
            # Right arc: center(15,19) r=4, 0° → -90° (east→north in screen coords)
            # Left  arc: center(9,19)  r=4, -90° → -180° (north→west in screen coords)
            def _arc(cx_, cy_, r_, a0, a1, n=10):
                return [(cx_ + r_*math.cos(a0 + (a1-a0)*i/(n-1)),
                         cy_ + r_*math.sin(a0 + (a1-a0)*i/(n-1)))
                        for i in range(n)]
            right_arc = _arc(15, 19, 4,  0,             -math.pi/2)
            left_arc  = _arc( 9, 19, 4, -math.pi/2, -math.pi)
            pts = ([(19, 21), (19, 19)]   # M19,21 v-2
                   + right_arc[1:]         # arc to (15,15)
                   + [(9, 15)]             # H9
                   + left_arc[1:]          # arc to (5,19)
                   + [(5, 21)])            # v2
            flat = [v for x_, y_ in pts for v in (x_*_s+ox, y_*_s+oy)]
            c.create_line(*flat, **lkw)

        def _icon_chart_network(c, ox, oy, color):
            """Lucide 'chart-network': axes + 3 node dots + connecting lines."""
            lkw = dict(fill=color, width=ICON_LW, capstyle="round", joinstyle="round")
            # L-shaped axes: vertical (3,3)→(3,19), corner arc, horizontal (5,21)→(21,21)
            c.create_line(3*_s+ox, 3*_s+oy, 3*_s+ox, 19*_s+oy, **lkw)
            cpts = []
            for i in range(11):
                a = math.radians(180 + 90*i/10)
                cpts += [5*_s+ox + 2*_s*math.cos(a), 19*_s+oy - 2*_s*math.sin(a)]
            c.create_line(*cpts, **lkw)
            c.create_line(5*_s+ox, 21*_s+oy, 21*_s+ox, 21*_s+oy, **lkw)
            # Connecting lines between nodes (boundary-to-boundary as in SVG)
            c.create_line(13.11*_s+ox, 7.664*_s+oy, 14.89*_s+ox, 10.336*_s+oy, **lkw)
            c.create_line(14.162*_s+ox, 12.788*_s+oy, 10.838*_s+ox, 14.212*_s+oy, **lkw)
            c.create_line(20*_s+ox, 4*_s+oy, 13.94*_s+ox, 5.515*_s+oy, **lkw)
            # Node circles (r=2) — same approach as logo badge dots
            cr = 2*_s
            ckw = dict(outline=color, fill="", width=ICON_LW)
            for nx, ny in [(12,6), (16,12), (9,15)]:
                c.create_oval(nx*_s+ox-cr, ny*_s+oy-cr, nx*_s+ox+cr, ny*_s+oy+cr, **ckw)

        def _draw_toggle(active_dm: bool):
            toggle_c.delete("all")
            _fill_pill(toggle_c, 0, 0, PILL_W, PILL_H, TRACK_BG)
            tp = THUMB_PAD
            if active_dm:
                _fill_pill(toggle_c, tp, tp, SEG_W - tp, PILL_H - tp, THUMB_BG)
            else:
                _fill_pill(toggle_c, SEG_W + tp, tp, PILL_W - tp, PILL_H - tp, THUMB_BG)
            dm_col  = ACT_FG if active_dm else INACT_FG
            an_col  = INACT_FG if active_dm else ACT_FG
            icon_oy = (PILL_H - ICON_SZ) / 2
            _icon_circle_user(toggle_c,   (SEG_W - ICON_SZ) / 2,         icon_oy, dm_col)
            _icon_chart_network(toggle_c, SEG_W + (SEG_W - ICON_SZ) / 2, icon_oy, an_col)

        def _set_nav_active(is_analysis: bool):
            _draw_toggle(not is_analysis)

        proj._nav_an_btn     = toggle_c
        proj._set_nav_active = _set_nav_active
        _draw_toggle(True)

        # ── Tooltip on hover ──────────────────────────────────────────────
        _tip     = [None]
        _tip_txt = [None]

        def _tip_show(text, event):
            if _tip_txt[0] == text and _tip[0] and _tip[0].winfo_exists():
                return
            _tip_hide()
            _tip_txt[0] = text
            tw = tk.Toplevel(toggle_c)
            tw.wm_overrideredirect(True)
            tw.configure(bg="#2c3b4e")
            tk.Label(tw, text=text, bg="#2c3b4e", fg="#ffffff",
                     font=(FONT_FAMILY, 10), padx=8, pady=4).pack()
            tw.update_idletasks()
            # Anchor below the pill; clamp so it never leaves the screen
            tip_w    = tw.winfo_reqwidth()
            tip_h    = tw.winfo_reqheight()
            scr_w    = toggle_c.winfo_screenwidth()
            scr_h    = toggle_c.winfo_screenheight()
            x = toggle_c.winfo_rootx() + (0 if event.x < SEG_W else SEG_W)
            y = toggle_c.winfo_rooty() + PILL_H + 6
            x = max(0, min(x, scr_w - tip_w - 4))
            y = max(0, min(y, scr_h - tip_h - 4))
            tw.wm_geometry(f"+{x}+{y}")
            _tip[0] = tw

        def _tip_hide(event=None):
            if _tip[0] and _tip[0].winfo_exists():
                _tip[0].destroy()
            _tip[0]     = None
            _tip_txt[0] = None

        def _on_motion(event):
            _tip_show("Decision Makers" if event.x < SEG_W else "Analysis", event)

        def _on_toggle_click(event):
            if event.x < SEG_W:
                self._show_matrix_view(proj)
            else:
                if proj.agg_tab_ids:
                    self._show_analysis_view(proj)

        toggle_c.bind("<Motion>",    _on_motion)
        toggle_c.bind("<Leave>",     _tip_hide)
        toggle_c.bind("<Button-1>",  _on_toggle_click)

        # ── Matrix view (default) ─────────────────────────────────────────
        matrix_frame = tk.Frame(outer, bg=COLOR_BG)
        matrix_frame.pack(fill="both", expand=True)
        proj._matrix_frame = matrix_frame

        inner_nb = ttk.Notebook(matrix_frame, style="Headless.TNotebook")
        inner_nb.pack(fill="both", expand=True)
        proj.notebook = inner_nb
        inner_nb.bind("<<NotebookTabChanged>>",
                      lambda e, p=proj: (self._update_topbar(),
                                         self._refresh_dm_underlines(p),
                                         self._ensure_current_dm_widget(p)))

        empty = tk.Label(
            matrix_frame,
            text='Click  "+ Add Decision-Maker"  to add the first matrix.',
            font=(FONT_FAMILY, FONT_SIZE_HEADER),
            bg=COLOR_BG, fg=COLOR_TEXT_LIGHT,
        )
        empty.place(relx=0.5, rely=0.55, anchor="center")
        proj._empty_label = empty

        # ── Analysis view (hidden until Run Analysis) ─────────────────────
        analysis_frame = tk.Frame(outer, bg=COLOR_BG)
        analysis_frame.pack_propagate(False)
        proj._analysis_frame  = analysis_frame
        proj._in_analysis_view = False

        # Breadcrumb bar removed — navigation is now handled by the switcher bar above
        crumb_bar = tk.Frame(analysis_frame, bg="#ffffff", height=0)
        crumb_bar.pack(fill="x")
        proj._crumb_bar = crumb_bar

        # Analysis notebook
        analysis_nb = ttk.Notebook(analysis_frame, style="Headless.TNotebook")
        analysis_nb.pack(fill="both", expand=True)
        proj._analysis_nb = analysis_nb
        analysis_nb.bind("<<NotebookTabChanged>>",
                         lambda e, p=proj: (self._ensure_selected_analysis_tab_loaded(p),
                                            self._update_analysis_breadcrumb(p),
                                            self._refresh_an_underlines(p)))

        self.projects.append(proj)
        self._sidebar_open_states[name] = True
        self._refresh_sidebar_projects()
        self._set_status(f'Project "{name}" created.')
        self._update_topbar()
        return proj

    # ------------------------------------------------------------------
    # Sidebar project list helpers
    # ------------------------------------------------------------------

    def _make_chevron(self, parent, direction="right"):
        """Canvas drawing of a chevron-right or chevron-down icon (16×16)."""
        size = 16
        c = tk.Canvas(parent, width=size, height=size,
                      bg="#fafbfc", highlightthickness=0)
        s = size / 24.0
        if direction == "right":
            pts = [9*s, 18*s, 15*s, 12*s, 9*s, 6*s]
        else:
            pts = [6*s, 9*s, 12*s, 15*s, 18*s, 9*s]
        c.create_line(*pts, fill="#a3a3a3", width=1.35,
                      capstyle="round", joinstyle="round")
        return c

    def _make_plus(self, parent, bg="#fafbfc"):
        """Canvas drawing of a plus icon (16×16)."""
        size = 16
        c = tk.Canvas(parent, width=size, height=size,
                      bg=bg, highlightthickness=0)
        s = size / 24.0
        c.create_line(5*s, 12*s, 19*s, 12*s,
                      fill="#a3a3a3", width=1.35, capstyle="round")
        c.create_line(12*s, 5*s, 12*s, 19*s,
                      fill="#a3a3a3", width=1.35, capstyle="round")
        return c

    def _make_user_icon(self, parent, bg="#fafbfc"):
        """Canvas drawing of Lucide user icon (16×16)."""
        size = 16
        c = tk.Canvas(parent, width=size, height=size,
                      bg=bg, highlightthickness=0)
        s = size / 24.0
        # Body arc: M19 21v-2a4 4 0 0 0-4-4H9a4 4 0 0 0-4 4v2
        # Approximate shoulders with a polyline
        c.create_line(
            19*s, 21*s, 19*s, 19*s,
            fill="#a3a3a3", width=1.35, capstyle="round", joinstyle="round",
        )
        c.create_line(
            5*s, 21*s, 5*s, 19*s,
            fill="#a3a3a3", width=1.35, capstyle="round", joinstyle="round",
        )
        # Shoulder curve approximated as arc
        c.create_arc(5*s, 11*s, 19*s, 23*s,
                     start=0, extent=180,
                     outline="#a3a3a3", width=1.35, style="arc")
        # Head circle: cx=12, cy=7, r=4
        r = 4*s
        cx, cy = 12*s, 7*s
        c.create_oval(cx-r, cy-r, cx+r, cy+r,
                      outline="#a3a3a3", width=1.35, fill="")
        return c

    def _make_folder(self, parent):
        """Canvas drawing of folder icon tracing the Lucide SVG path (16×16)."""
        size = 16
        c = tk.Canvas(parent, width=size, height=size,
                      bg="#fafbfc", highlightthickness=0)
        s = size / 24.0
        # Trace: M20 20 ... a2 2 ... V8 ... h-7.9 ... L9.6 3.9 ... H4 ... v13 ... Z
        # Approximated with straight-line segments (corner radii are ~1px at this scale)
        pts = [
            2*s,    5*s,    # top-left (after corner)
            4*s,    3*s,    # top-left body start
            7.93*s, 3*s,    # tab top-left
            9.6*s,  3.9*s,  # tab curve point
            12.1*s, 6*s,    # tab right meets body
            20*s,   6*s,    # top-right (before corner)
            22*s,   8*s,    # top-right (after corner)
            22*s,   18*s,   # bottom-right (before corner)
            20*s,   20*s,   # bottom-right (after corner)
            2*s,    20*s,   # bottom-left (before corner)
            2*s,    5*s,    # close back to start
        ]
        c.create_line(pts, fill="#a3a3a3", width=1.35,
                      capstyle="round", joinstyle="round")
        return c

    def _refresh_sidebar_projects(self):
        """Rebuild the project rows in the sidebar, preserving open/close state."""
        if not hasattr(self, "_sidebar_open_states"):
            self._sidebar_open_states = {}

        for w in self._sidebar_proj_list.winfo_children():
            w.destroy()

        # ── Collapsed view ─────────────────────────────────────────────
        if not getattr(self, "_sidebar_expanded", True):
            for proj in self.projects:
                abbrev = proj.name[:2].upper()
                _BG    = "#eaeef4"
                _FG    = "#30455c"
                _R     = 4
                _f     = tkFont.Font(family=FONT_FAMILY, size=12, weight="bold")
                bw = 37
                bh = 37
                wrap = tk.Frame(self._sidebar_proj_list, bg="#fafbfc")
                wrap.pack(fill="x", pady=4)
                bc = tk.Canvas(wrap, width=bw, height=bh,
                               bg="#fafbfc", highlightthickness=0, cursor=CURSOR_HAND)
                bc.pack()
                r = _R; d = r * 2
                x1, y1, x2, y2 = 1, 1, bw-1, bh-1
                kw = dict(fill=_BG, outline=_BG)
                bc.create_arc(x1,    y1,    x1+d, y1+d, start=90,  extent=90, **kw)
                bc.create_arc(x2-d,  y1,    x2,   y1+d, start=0,   extent=90, **kw)
                bc.create_arc(x1,    y2-d,  x1+d, y2,   start=180, extent=90, **kw)
                bc.create_arc(x2-d,  y2-d,  x2,   y2,   start=270, extent=90, **kw)
                bc.create_rectangle(x1+r, y1, x2-r, y2, fill=_BG, outline=_BG)
                bc.create_rectangle(x1, y1+r, x2, y2-r, fill=_BG, outline=_BG)
                bc.create_text(bw // 2, bh // 2, text=abbrev,
                               fill=_FG, font=_f, anchor="center")
                bc.bind("<Button-1>", lambda e=None, p=proj: self._select_project(p))
            return

        # ── Expanded view ──────────────────────────────────────────────
        if not self.projects:
            tk.Label(
                self._sidebar_proj_list,
                text="No projects available.\nCreate a new project to begin assessing interactions between policies across decision-makers.",
                font=(FONT_FAMILY, 11),
                bg="#fafbfc", fg="#a3a3a3",
                justify="left", anchor="w",
                wraplength=270,
            ).pack(anchor="w", padx=20, pady=(4, 0))
            return

        _HOVER_BG  = "#ebebeb"
        _NORMAL_BG = "#fafbfc"

        for proj in self.projects:
            is_open = self._sidebar_open_states.get(proj.name, False)

            row = tk.Frame(self._sidebar_proj_list, bg="#fafbfc")
            row.pack(fill="x", padx=20, pady=(12, 0))

            chevron = self._make_chevron(row, "down" if is_open else "right")
            chevron.config(cursor=CURSOR_HAND)
            chevron.grid(row=0, column=0, padx=(0, 4), pady=(3, 0), sticky="n")

            folder = self._make_folder(row)
            folder.config(cursor=CURSOR_HAND)
            folder.grid(row=0, column=1, padx=(0, 6), pady=(3, 0), sticky="n")

            name_lbl = tk.Label(
                row, text=proj.name,
                font=(FONT_FAMILY, 14),
                bg="#fafbfc", fg="#2c3b4e", anchor="nw",
                justify="left", wraplength=236,
                pady=0, cursor=CURSOR_HAND,
            )
            name_lbl.grid(row=0, column=2, sticky="nw")
            row.grid_columnconfigure(2, weight=1)

            # Collapsible content
            content = tk.Frame(self._sidebar_proj_list, bg="#fafbfc")
            tk.Frame(content, bg="#fafbfc", height=8).pack(fill="x")  # top spacer

            # Add DM button — canvas for rounded corners
            dm_wrap = tk.Frame(content, bg="#fafbfc")
            dm_wrap.pack(fill="x", padx=20, pady=(0, 2))

            _DM_RR       = 5
            _dm_btn_t    = [0.0]
            _dm_btn_anim = [None]
            _dm_bg_items = []

            dm_row = tk.Canvas(dm_wrap, height=32, bg="#fafbfc",
                               highlightthickness=0, cursor=CURSOR_HAND)
            dm_row.pack(fill="x")

            dm_lbl = tk.Label(
                dm_row, text="DECISION-MAKERS",
                font=(FONT_FAMILY, 11),
                bg=_NORMAL_BG, fg="#a3a3a3", anchor="w",
                padx=0, pady=0,
            )
            lbl_win  = dm_row.create_window(20, 15, anchor="w", window=dm_lbl)
            plus_icon = self._make_plus(dm_row, bg=_NORMAL_BG)
            icon_win = [dm_row.create_window(0, 15, anchor="e", window=plus_icon)]

            def _dm_rebuild_rr(color, _cv=dm_row, _bi=_dm_bg_items, _lw=lbl_win,
                               _iw=icon_win, _lbl=dm_lbl, _pi=plus_icon):
                for item in _bi:
                    try: _cv.delete(item)
                    except tk.TclError: pass
                _bi.clear()
                W = _cv.winfo_width()
                if W < 2:
                    return
                H = 30; r = _DM_RR; d = r * 2
                x1, y1, x2, y2 = 1, 1, W-1, H-1
                kw = dict(fill=color, outline=color)
                _bi.extend([
                    _cv.create_arc(x1,    y1,    x1+d, y1+d, start=90,  extent=90, **kw),
                    _cv.create_arc(x2-d,  y1,    x2,   y1+d, start=0,   extent=90, **kw),
                    _cv.create_arc(x1,    y2-d,  x1+d, y2,   start=180, extent=90, **kw),
                    _cv.create_arc(x2-d,  y2-d,  x2,   y2,   start=270, extent=90, **kw),
                    _cv.create_rectangle(x1+r, y1,   x2-r, y2,   **kw),
                    _cv.create_rectangle(x1,   y1+r, x2,   y2-r, **kw),
                ])
                _cv.tag_raise(_lw)
                _cv.tag_raise(_iw[0])
                _cv.coords(_iw[0], W - 12, 15)
                try: _lbl.config(bg=color); _pi.config(bg=color)
                except tk.TclError: pass

            def _dm_update_color(color, _cv=dm_row, _bi=_dm_bg_items,
                                 _lbl=dm_lbl, _pi=plus_icon):
                for item in _bi:
                    try: _cv.itemconfig(item, fill=color, outline=color)
                    except tk.TclError: pass
                try: _lbl.config(bg=color); _pi.config(bg=color)
                except tk.TclError: pass

            def _dm_btn_animate(target, _cv=dm_row, _t=_dm_btn_t,
                                _anim=_dm_btn_anim, _upd=_dm_update_color):
                if _anim[0]:
                    _cv.after_cancel(_anim[0])
                    _anim[0] = None
                def tick():
                    diff = target - _t[0]
                    if abs(diff) < 0.02:
                        _t[0] = target
                        _upd(_hex_interp(_NORMAL_BG, _HOVER_BG, target))
                        _anim[0] = None
                        return
                    _t[0] += diff * 0.3
                    _upd(_hex_interp(_NORMAL_BG, _HOVER_BG, _t[0]))
                    _anim[0] = _cv.after(16, tick)
                tick()

            def _hover_in(e, animate=_dm_btn_animate):
                animate(1.0)

            def _hover_out(e, animate=_dm_btn_animate):
                animate(0.0)

            dm_row.bind("<Configure>", lambda e=None, rb=_dm_rebuild_rr, t=_dm_btn_t: rb(
                _hex_interp(_NORMAL_BG, _HOVER_BG, t[0])))
            dm_row.after(50, lambda rb=_dm_rebuild_rr: rb(_NORMAL_BG))

            for w in (dm_wrap, dm_row, dm_lbl, plus_icon):
                w.bind("<Enter>", _hover_in)
                w.bind("<Leave>", _hover_out)
                w.bind("<Button-1>", lambda e=None, p=proj: self._add_matrix(p))

            # DM list
            dm_list_frame = tk.Frame(content, bg="#fafbfc")
            dm_list_frame.pack(fill="x", pady=(1, 0))

            _DM_ITEM_RR    = 5
            _DM_ITEM_HOVER = "#eaeef4"

            for matrix in proj.matrices:
                dm_item_wrap = tk.Frame(dm_list_frame, bg="#fafbfc")
                dm_item_wrap.pack(fill="x", padx=20, pady=1)

                _item_t       = [0.0]
                _item_anim    = [None]
                _item_bg_itms = []

                dm_item = tk.Canvas(dm_item_wrap, height=36, bg="#fafbfc",
                                    highlightthickness=0, cursor=CURSOR_HAND)
                dm_item.pack(fill="x")

                user_ic = self._make_user_icon(dm_item, bg="#fafbfc")
                u_win = dm_item.create_window(28, 18, anchor="w", window=user_ic)

                dm_name_lbl = tk.Label(
                    dm_item, text=matrix.decision_maker,
                    font=(FONT_FAMILY, 13),
                    bg="#fafbfc", fg="#426387", anchor="w",
                    justify="left", wraplength=210,
                    pady=0,
                )
                n_win = dm_item.create_window(50, 18, anchor="w", window=dm_name_lbl)

                def _resize_canvas(e, cv=dm_item, lbl=dm_name_lbl,
                                   uw=u_win, nw=n_win):
                    lh = lbl.winfo_reqheight()
                    new_h = max(36, lh + 16)
                    cv.config(height=new_h)
                    cy = new_h // 2
                    cv.coords(uw, 28, cy)
                    cv.coords(nw, 50, cy)

                dm_name_lbl.bind("<Configure>", _resize_canvas)

                def _item_rebuild_rr(color, cv=dm_item, bi=_item_bg_itms,
                                     uw=u_win, nw=n_win,
                                     u=user_ic, lbl=dm_name_lbl):
                    for item in bi:
                        try: cv.delete(item)
                        except tk.TclError: pass
                    bi.clear()
                    W = cv.winfo_width()
                    H = cv.winfo_height()
                    if W < 2 or H < 2:
                        return
                    r = _DM_ITEM_RR; d = r * 2
                    x1, y1, x2, y2 = 1, 1, W-1, H-1
                    kw = dict(fill=color, outline=color)
                    bi.extend([
                        cv.create_arc(x1,    y1,    x1+d, y1+d, start=90,  extent=90, **kw),
                        cv.create_arc(x2-d,  y1,    x2,   y1+d, start=0,   extent=90, **kw),
                        cv.create_arc(x1,    y2-d,  x1+d, y2,   start=180, extent=90, **kw),
                        cv.create_arc(x2-d,  y2-d,  x2,   y2,   start=270, extent=90, **kw),
                        cv.create_rectangle(x1+r, y1,   x2-r, y2,   **kw),
                        cv.create_rectangle(x1,   y1+r, x2,   y2-r, **kw),
                    ])
                    cv.tag_raise(uw)
                    cv.tag_raise(nw)
                    try: u.config(bg=color); lbl.config(bg=color)
                    except tk.TclError: pass

                def _item_update_color(color, cv=dm_item, bi=_item_bg_itms,
                                       u=user_ic, lbl=dm_name_lbl):
                    for item in bi:
                        try: cv.itemconfig(item, fill=color, outline=color)
                        except tk.TclError: pass
                    try: u.config(bg=color); lbl.config(bg=color)
                    except tk.TclError: pass

                def _dm_item_animate(target, cv=dm_item, t=_item_t, anim=_item_anim,
                                     upd=_item_update_color):
                    if anim[0]:
                        cv.after_cancel(anim[0])
                        anim[0] = None
                    def tick():
                        diff = target - t[0]
                        if abs(diff) < 0.02:
                            t[0] = target
                            upd(_hex_interp(_NORMAL_BG, _DM_ITEM_HOVER, target))
                            anim[0] = None
                            return
                        t[0] += diff * 0.3
                        upd(_hex_interp(_NORMAL_BG, _DM_ITEM_HOVER, t[0]))
                        anim[0] = cv.after(16, tick)
                    tick()

                def _dm_hover_in(e, animate=_dm_item_animate):
                    animate(1.0)

                def _dm_hover_out(e, animate=_dm_item_animate):
                    animate(0.0)

                def _dm_click(e, p=proj, m=matrix):
                    self._select_project(p)
                    if hasattr(m, "_tab"):
                        p.notebook.select(m._tab)

                dm_item.bind("<Configure>", lambda e=None, rb=_item_rebuild_rr, t=_item_t:
                             rb(_hex_interp(_NORMAL_BG, _DM_ITEM_HOVER, t[0])))
                dm_item.after(50, lambda rb=_item_rebuild_rr: rb(_NORMAL_BG))

                for w in (dm_item_wrap, dm_item, dm_name_lbl, user_ic):
                    w.bind("<Enter>", _dm_hover_in)
                    w.bind("<Leave>", _dm_hover_out)
                    w.bind("<Button-1>", _dm_click)

            # Bottom spacer for the project block
            tk.Frame(content, bg="#fafbfc", height=10).pack(fill="x")

            def _toggle(event=None, pname=proj.name):
                self._sidebar_open_states[pname] = not self._sidebar_open_states.get(pname, False)
                self._refresh_sidebar_projects()

            if is_open:
                content.pack(fill="x", after=row)

            def _navigate(e=None, p=proj):
                self._select_project(p)

            chevron.bind("<Button-1>", _toggle)
            name_lbl.bind("<Button-1>", _navigate)
            folder.bind("<Button-1>",   _navigate)

        # Bind mousewheel on all sidebar list widgets so scrolling works everywhere
        def _bind_scroll_recursive(w):
            c = getattr(self, "_sidebar_proj_canvas", None)
            if not c:
                return
            w.bind("<MouseWheel>",
                   lambda e=None, _c=c: _c.yview_scroll(-1 if getattr(e, "delta", 0) > 0 else 1, "units"))
            w.bind("<Button-4>",
                   lambda e=None, _c=c: _c.yview_scroll(-1, "units"))
            w.bind("<Button-5>",
                   lambda e=None, _c=c: _c.yview_scroll( 1, "units"))
            for ch in w.winfo_children():
                _bind_scroll_recursive(ch)
        self.root.after_idle(lambda: _bind_scroll_recursive(self._sidebar_proj_list))

    def _current_project(self) -> Optional[Project]:
        """Return the currently selected project, or None."""
        if not self._proj_nb.tabs():
            return None
        try:
            idx = self._proj_nb.index("current")
            return self.projects[idx]
        except (tk.TclError, IndexError):
            return None

    def _rename_project(self):
        proj = self._current_project()
        if not proj:
            messagebox.showinfo("No Project", "No project is open.", parent=self.root)
            return
        existing = [p.name for p in self.projects if p is not proj]
        dlg = _SimpleInputDialog(self.root, "Rename Project",
                                 "New project name:", proj.name)
        self.root.wait_window(dlg)
        if dlg.result and dlg.result != proj.name:
            if dlg.result in [p.name for p in self.projects]:
                messagebox.showwarning("Duplicate Name",
                                       f'A project named "{dlg.result}" already exists.',
                                       parent=self.root)
                return
            idx = self._proj_nb.index("current")
            proj.name = dlg.result
            self._proj_nb.tab(idx, text=f"  {dlg.result}  ")
            # Update project header bar
            for widget in proj.frame.winfo_children():
                if isinstance(widget, tk.Frame) and widget.cget("bg") == COLOR_ACCENT:
                    for lbl in widget.winfo_children():
                        if isinstance(lbl, tk.Label):
                            lbl.config(text=f"Project:  {dlg.result}")
            self._set_status(f'Project renamed to "{dlg.result}".')

    def _delete_project(self):
        proj = self._current_project()
        if not proj:
            messagebox.showinfo("No Project", "No project is open.",
                                parent=self.root)
            return
        if not messagebox.askyesno(
            "Delete Project",
            f'Permanently delete project "{proj.name}" and all its data?\n'
            "This cannot be undone.",
            parent=self.root,
        ):
            return
        idx = self._proj_nb.index("current")
        self._proj_nb.forget(idx)
        self.projects.pop(idx)
        self._refresh_sidebar_projects()
        if not self.projects:
            self._empty_label.place(x=0, y=0, relwidth=1, relheight=1)
        self._set_status(f'Project "{proj.name}" deleted.')

    # ==================================================================
    # Decision-maker tab management (operates on current project)
    # ==================================================================

    def _add_matrix(self, proj=None):
        if proj is None:
            proj = self._current_project()
        if not proj:
            messagebox.showinfo("No Project",
                                "Create a project first.", parent=self.root)
            return

        existing_dm_names = [m.decision_maker for m in proj.matrices]
        dlg = ProjectSetupDialog(
            self.root,
            title="Add Decision-Makers",
            include_project_name=False,
            existing_dm_names=existing_dm_names,
            fixed_policies=proj.matrices[0].policies if proj.matrices else None,
        )
        self.root.wait_window(dlg)
        if dlg.result is None:
            return

        data = dlg.result
        if data.get("action") == "import_excel":
            self._import_excel_into_project(proj)
            return
        policies = proj.matrices[0].policies if proj.matrices else data["policies"]

        for dm_name in data["decision_makers"]:
            matrix = make_empty_matrix(dm_name, policies)
            proj.matrices.append(matrix)
            self._create_dm_tab(proj, matrix)
        self._proj_nb.select(proj.frame)
        self._refresh_sidebar_projects()
        self._set_status(
            f'Added {len(data["decision_makers"])} decision-maker'
            f'{"s" if len(data["decision_makers"]) != 1 else ""} to project "{proj.name}".'
        )
        self._update_topbar()

    def _create_dm_tab(self, proj: Project, matrix: PolicyMatrix):
        """Build one DM matrix tab inside the project's inner notebook."""
        proj._empty_label.place_forget()

        tab = tk.Frame(proj.notebook, bg=COLOR_BG)
        proj.notebook.add(tab, text=matrix.decision_maker)
        proj.notebook.select(tab)
        matrix._tab = tab
        tab._dm_matrix = matrix   # stored for lazy widget construction
        tab._dm_widget_built = False
        self._add_dm_tab_label(proj, matrix, tab)

    def _ensure_current_dm_widget(self, proj: Project):
        """Lazily build the MatrixWidget for the currently selected DM tab."""
        try:
            selected = proj.notebook.select()
        except (tk.TclError, AttributeError):
            return
        for tab_id in proj.notebook.tabs():
            if str(tab_id) != str(selected):
                continue
            tab = proj.notebook.nametowidget(tab_id)
            if not getattr(tab, "_dm_widget_built", True):
                matrix = tab._dm_matrix

                def on_change(r, c, v, m=matrix):
                    filled = m.filled_count()
                    total  = m.total_cells()
                    self._set_status(
                        f'"{m.decision_maker}"  --  {filled}/{total} cells filled  '
                        f'|  {m.codes[r]} -> {m.codes[c]}: {v}'
                    )

                mw = MatrixWidget(tab, matrix, on_change=on_change)
                mw.pack(fill="both", expand=True, padx=8, pady=8)
                tab._dm_widget_built = True
            break

    def _current_inner_index(self, proj: Project) -> int:
        if not proj.notebook.tabs():
            return -1
        try:
            return proj.notebook.index("current")
        except tk.TclError:
            return -1

    def _rename_current_tab(self):
        proj = self._current_project()
        if not proj:
            return
        if getattr(proj, "_in_analysis_view", False):
            messagebox.showinfo("Cannot Rename",
                                "Analysis result tabs cannot be renamed.",
                                parent=self.root)
            return
        idx = self._current_inner_index(proj)
        if idx < 0:
            return
        matrix = proj.matrices[idx]
        dlg = _SimpleInputDialog(self.root, "Rename Tab",
                                 "New decision-maker name:", matrix.decision_maker)
        self.root.wait_window(dlg)
        if dlg.result:
            matrix.decision_maker = dlg.result
            proj.notebook.tab(idx, text=dlg.result)
            if idx < len(proj._dm_tab_entries):
                proj._dm_tab_entries[idx][1].configure(text=dlg.result)
            self._set_status(f'Renamed to "{dlg.result}".')

    def _remove_current_tab(self):
        proj = self._current_project()
        if not proj:
            return

        if getattr(proj, "_in_analysis_view", False):
            # Remove selected analysis tab
            if not messagebox.askyesno("Remove Tab",
                                       "Remove this analysis result tab?",
                                       parent=self.root):
                return
            try:
                idx    = proj._analysis_nb.index("current")
                tab_id = proj._analysis_nb.tabs()[idx]
                proj._analysis_nb.forget(idx)
                if tab_id in proj.agg_tab_ids:
                    proj.agg_tab_ids.remove(tab_id)
                proj.analysis_tab_meta.pop(tab_id, None)
                for i, (tid, _, _, cont) in enumerate(proj._an_tab_entries):
                    if str(tid) == str(tab_id):
                        cont.destroy()
                        proj._an_tab_entries.pop(i)
                        break
                self._refresh_an_underlines(proj)
            except tk.TclError:
                pass
            self._set_status("Analysis tab removed.")
        else:
            # Remove selected DM tab
            idx = self._current_inner_index(proj)
            if idx < 0:
                return
            dm = proj.matrices[idx].decision_maker
            if not messagebox.askyesno("Remove Matrix",
                                       f'Remove matrix for "{dm}"?',
                                       parent=self.root):
                return
            proj.notebook.forget(idx)
            proj.matrices.pop(idx)
            if idx < len(proj._dm_tab_entries):
                _, _, _, cont = proj._dm_tab_entries.pop(idx)
                cont.destroy()
            self._refresh_dm_underlines(proj)
            if not proj.matrices:
                proj._empty_label.place(relx=0.5, rely=0.55, anchor="center")
            self._set_status(f'Removed "{dm}".')
            self._update_topbar()

    # ==================================================================
    # Aggregation / Analysis (operates on current project)
    # ==================================================================

    def _run_aggregation(self):
        proj = self._current_project()
        if not proj:
            messagebox.showinfo("No Project", "Create a project first.",
                                parent=self.root)
            return

        if len(proj.matrices) < 2:
            messagebox.showwarning("Not Enough Matrices",
                                   "You need at least 2 decision-maker matrices.",
                                   parent=self.root)
            return

        incomplete = check_completeness(proj.matrices)
        if incomplete:
            lines = []
            for dm_name, blanks in incomplete:
                cells = ", ".join(f"{r}->{c}" for r, c in blanks[:10])
                suffix = f" (+{len(blanks)-10} more)" if len(blanks) > 10 else ""
                lines.append(f"  {dm_name}: {cells}{suffix}")
            messagebox.showerror("Incomplete Matrices",
                                 "Please fill all cells before aggregating.\n\n"
                                 + "\n".join(lines), parent=self.root)
            return

        n_dms = len(proj.matrices)
        if n_dms < 3:
            method = "average"
        else:
            dlg = AggregationMethodDialog(
                self.root, [m.decision_maker for m in proj.matrices])
            self.root.wait_window(dlg)
            if dlg.result is None:
                return
            method = dlg.result

        weights = None
        if method == "weighted":
            dlg_w = WeightDialog(
                self.root, [m.decision_maker for m in proj.matrices])
            self.root.wait_window(dlg_w)
            if dlg_w.result is None:
                return
            weights = dlg_w.result

        if method == "average":
            result = aggregate_average(proj.matrices)
        elif method == "majority":
            result = aggregate_majority(proj.matrices)
        elif method == "weighted":
            result = aggregate_weighted(proj.matrices, weights)
        else:
            return

        if result.ties:
            dlg_t = TieResolutionDialog(self.root, result.ties)
            self.root.wait_window(dlg_t)
            if dlg_t.result is None:
                return
            resolve_ties(result)

        self._create_analysis_tabs(proj, result)
        self._set_status(
            f'Analysis complete for "{proj.name}"  |  method: {method}  '
            f'|  {n_dms} decision-makers'
        )

    def _create_analysis_tabs(self, proj: Project, result: AggregationResult):
        """Populate the analysis notebook and switch to the analysis view."""
        method_label = {
            "average":  "Average",
            "majority": "Majority",
            "weighted": "Weighted",
        }.get(result.method, result.method.title())

        proj.agg_results.append(result)

        # Clear any existing analysis tabs
        for tab_id in list(proj.agg_tab_ids):
            try:
                proj._analysis_nb.forget(tab_id)
            except tk.TclError:
                pass
        proj.agg_tab_ids.clear()
        proj.analysis_tab_meta.clear()

        tabs = [
            (f"  Results Insights ({method_label})  ",  ResultsInsightsTab),
            (f"  Aggregated ({method_label})  ",         AggregationTab),
            (f"  Coherence Scores ({method_label})  ",   CoherenceScoresTab),
            (f"  Range of Influence ({method_label})  ", RangeOfInfluenceTab),
            (f"  PCA ({method_label})  ",                PCATab),
            (f"  Network Analysis ({method_label})  ",   NetworkTab),
            (f"  LLM Interpretation ({method_label})  ", LLMInterpretationTab),
        ]

        first_tab_id = None
        for title, TabClass in tabs:
            host = tk.Frame(proj._analysis_nb, bg=COLOR_BG)
            proj._analysis_nb.add(host, text=title)
            tab_id = proj._analysis_nb.tabs()[-1]
            proj.agg_tab_ids.append(tab_id)
            proj.analysis_tab_meta[tab_id] = {
                "host": host,
                "tab_class": TabClass,
                "result": result,
                "loaded": False,
            }
            if first_tab_id is None:
                first_tab_id = tab_id

        if first_tab_id is not None:
            self._ensure_analysis_tab_loaded(proj, first_tab_id)
            proj._analysis_nb.select(first_tab_id)

        self._build_analysis_breadcrumb(proj)
        self._show_analysis_view(proj)

    def _ensure_selected_analysis_tab_loaded(self, proj: Project):
        try:
            tab_id = proj._analysis_nb.select()
        except (tk.TclError, AttributeError):
            return
        if tab_id:
            self._ensure_analysis_tab_loaded(proj, tab_id)

    def _ensure_analysis_tab_loaded(self, proj: Project, tab_id: str):
        meta = proj.analysis_tab_meta.get(tab_id)
        if not meta or meta.get("loaded"):
            return

        host = meta["host"]
        tab_class = meta["tab_class"]
        result = meta["result"]

        widget = tab_class(host, result)
        widget.pack(fill="both", expand=True)

        meta["widget"] = widget
        meta["loaded"] = True

    # ==================================================================
    # Analysis view navigation
    # ==================================================================

    # ------------------------------------------------------------------
    # Custom tab-strip helpers (DM labels + Analysis labels in switcher)
    # ------------------------------------------------------------------

    _TAB_BG          = "#ffffff"
    _TAB_ACTIVE_FG   = "#30455c"
    _TAB_INACTIVE_FG = "#a3a3a3"
    _TAB_UNDERLINE   = "#557ca2"
    _TAB_TEXT_SIZE   = 12          # pt — tab label font size
    _TAB_PADX        = 14          # horizontal inner padding on each label
    _TAB_UL_EXTRA    = 4           # px beyond text width on each side of underline

    def _bind_ul_to_text(self, lbl, ul):
        """After each layout pass, resize ul so it is _TAB_UL_EXTRA px beyond the text on each side."""
        import tkinter.font as tkFont

        def _update(e):
            try:
                font = tkFont.nametofont(lbl.cget("font"))
            except Exception:
                font = tkFont.Font(family=FONT_FAMILY, size=self._TAB_TEXT_SIZE)
            tw   = font.measure(lbl.cget("text"))
            padx = max(0, (e.width - tw) // 2 - self._TAB_UL_EXTRA)
            ul.pack_configure(padx=padx)

        lbl.bind("<Configure>", _update)

    def _add_dm_tab_label(self, proj, matrix, nb_tab):
        """Append one DM tab label to the switcher's dm_tab_bar."""
        container = tk.Frame(proj._dm_tab_bar, bg=self._TAB_BG, cursor=CURSOR_HAND)
        container.pack(side="left", fill="y")

        lbl = tk.Label(container, text=matrix.decision_maker,
                       bg=self._TAB_BG, fg=self._TAB_INACTIVE_FG,
                       padx=self._TAB_PADX, wraplength=0,
                       font=(FONT_FAMILY, self._TAB_TEXT_SIZE, "normal"))
        lbl.pack(side="top", fill="both", expand=True)

        ul = tk.Frame(container, height=1, bg=self._TAB_BG)
        ul.pack(side="bottom", fill="x", padx=self._TAB_PADX)
        ul.pack_propagate(False)
        self._bind_ul_to_text(lbl, ul)

        # Select on release so drag-scroll doesn't accidentally switch tabs
        _ds          = getattr(proj, "_dm_drag",        None)
        _drag_start  = getattr(proj, "_dm_drag_start",  None)
        _drag_motion = getattr(proj, "_dm_drag_motion", None)

        def _on_release(e, t=nb_tab, d=_ds):
            was_drag = d.get("dragging", False) if d else False
            if d:
                d["dragging"] = False
            if not was_drag:
                proj.notebook.select(t)

        for w in (container, lbl):
            if _drag_start:  w.bind("<Button-1>",        _drag_start)
            if _drag_motion: w.bind("<B1-Motion>",        _drag_motion, add="+")
            w.bind("<ButtonRelease-1>", _on_release)

        # Forward mousewheel to the horizontal scroll canvas
        _sc = getattr(proj, "_dm_scroll_c", None)
        _ds = getattr(proj, "_dm_drag", None)
        if _sc:
            def _hscroll(e, c=_sc):
                raw = getattr(e, "delta", 0)
                if not raw:
                    return
                steps = raw if sys.platform == "darwin" else raw / 120.0
                _animate_canvas_x_scroll(c, _ds, -steps * 140, duration=220)
            for w in (container, lbl, ul):
                w.bind("<MouseWheel>", _hscroll)

        proj._dm_tab_entries.append((nb_tab, lbl, ul, container))
        self._refresh_dm_underlines(proj)

    def _refresh_dm_underlines(self, proj):
        """Update active/inactive styling on all DM tab labels."""
        try:
            current = proj.notebook.select()
        except (tk.TclError, AttributeError):
            return
        for (nb_tab, lbl, ul, _) in proj._dm_tab_entries:
            active = (str(nb_tab) == str(current))
            lbl.configure(
                fg=self._TAB_ACTIVE_FG if active else self._TAB_INACTIVE_FG,
                font=(FONT_FAMILY, self._TAB_TEXT_SIZE, "normal"),
            )
            ul.configure(bg=self._TAB_UNDERLINE if active else self._TAB_BG)

    def _rebuild_an_tab_bar(self, proj):
        """Clear and rebuild the analysis tab strip from the current analysis notebook."""
        for (_, _, _, cont) in proj._an_tab_entries:
            cont.destroy()
        proj._an_tab_entries.clear()

        try:
            all_tabs = proj._analysis_nb.tabs()
        except (tk.TclError, AttributeError):
            return

        for tab_id in proj.agg_tab_ids:
            if tab_id not in all_tabs:
                continue
            title = proj._analysis_nb.tab(tab_id, "text").strip()

            container = tk.Frame(proj._an_tab_bar, bg=self._TAB_BG, cursor=CURSOR_HAND)
            container.pack(side="left", fill="y")

            lbl = tk.Label(container, text=title,
                           bg=self._TAB_BG, fg=self._TAB_INACTIVE_FG,
                           padx=self._TAB_PADX,
                           font=(FONT_FAMILY, FONT_SIZE_NORMAL, "normal"))
            lbl.pack(side="top", fill="both", expand=True)

            ul = tk.Frame(container, height=1, bg=self._TAB_BG)
            ul.pack(side="bottom", fill="x", padx=self._TAB_PADX)
            ul.pack_propagate(False)
            self._bind_ul_to_text(lbl, ul)

            _ds = getattr(proj, "_an_drag", None)
            _drag_start = getattr(proj, "_an_drag_start", None)
            _drag_motion = getattr(proj, "_an_drag_motion", None)

            def _on_release(e, tid=tab_id, d=_ds):
                was_drag = d.get("dragging", False) if d else False
                if d:
                    d["dragging"] = False
                if not was_drag:
                    proj._analysis_nb.select(tid)

            for w in (container, lbl):
                if _drag_start:
                    w.bind("<Button-1>", _drag_start)
                if _drag_motion:
                    w.bind("<B1-Motion>", _drag_motion, add="+")
                w.bind("<ButtonRelease-1>", _on_release)

            _sc = getattr(proj, "_an_scroll_c", None)
            _ds = getattr(proj, "_an_drag", None)
            if _sc:
                def _hscroll(e, c=_sc):
                    raw = getattr(e, "delta", 0)
                    if not raw:
                        return
                    steps = raw if sys.platform == "darwin" else raw / 120.0
                    _animate_canvas_x_scroll(c, _ds, -steps * 140, duration=220)
                for w in (container, lbl, ul):
                    w.bind("<MouseWheel>", _hscroll)

            proj._an_tab_entries.append((tab_id, lbl, ul, container))

        self._refresh_an_underlines(proj)

    def _refresh_an_underlines(self, proj):
        """Update active/inactive styling on all analysis tab labels."""
        try:
            current = proj._analysis_nb.select()
        except (tk.TclError, AttributeError):
            return
        for (tab_id, lbl, ul, _) in proj._an_tab_entries:
            active = (str(tab_id) == str(current))
            lbl.configure(
                fg=self._TAB_ACTIVE_FG if active else self._TAB_INACTIVE_FG,
                font=(FONT_FAMILY, self._TAB_TEXT_SIZE, "normal"),
            )
            ul.configure(bg=self._TAB_UNDERLINE if active else self._TAB_BG)

    def _show_analysis_view(self, proj: Project):
        if not proj.agg_tab_ids:
            return
        proj._matrix_frame.pack_forget()
        proj._dm_scroll_c.pack_forget()
        proj._analysis_frame.pack(fill="both", expand=True)
        proj._an_scroll_c.pack(side="left", fill="both", expand=True)
        proj._in_analysis_view = True
        proj._set_nav_active(True)
        self._refresh_an_underlines(proj)
        self._update_topbar()

    def _show_matrix_view(self, proj: Project):
        proj._analysis_frame.pack_forget()
        proj._an_scroll_c.pack_forget()
        proj._matrix_frame.pack(fill="both", expand=True)
        proj._dm_scroll_c.pack(side="left", fill="both", expand=True)
        proj._in_analysis_view = False
        proj._set_nav_active(False)
        self._refresh_dm_underlines(proj)
        self._update_topbar()

    def _build_analysis_breadcrumb(self, proj: Project):
        """Rebuild the analysis tab strip in the switcher bar."""
        self._rebuild_an_tab_bar(proj)
        self._update_analysis_breadcrumb(proj)

    def _update_analysis_breadcrumb(self, proj: Project):
        """Redraw toggle to reflect that analysis results now exist."""
        proj._set_nav_active(proj._in_analysis_view)

    # ==================================================================
    # Import (operates on current project)
    # ==================================================================

    def _import_excel(self):
        proj = self._current_project()
        if not proj:
            messagebox.showinfo("No Project", "Create a project first.",
                                parent=self.root)
            return
        self._import_excel_into_project(proj)

    def _ask_excel_path(self):
        return filedialog.askopenfilename(
            parent=self.root,
            title="Import matrices from Excel workbook",
            filetypes=[("Excel Workbook", "*.xlsx *.xls")],
        )

    def _load_import_result(self, path):
        try:
            return import_matrices_from_excel(path)
        except (ValueError, ImportError) as exc:
            messagebox.showerror("Import Error", str(exc), parent=self.root)
            return None

    def _unique_project_name(self, base_name: str) -> str:
        existing = {p.name for p in self.projects}
        if base_name not in existing:
            return base_name
        i = 2
        while f"{base_name} {i}" in existing:
            i += 1
        return f"{base_name} {i}"

    def _import_excel_into_project(self, proj: Project):
        path = self._ask_excel_path()
        if not path:
            return

        result = self._load_import_result(path)
        if result is None:
            return

        if not result.matrices:
            messagebox.showwarning("Nothing Imported",
                                   "No valid matrices were found in the workbook.",
                                   parent=self.root)
            return

        ref_policies = (
            proj.matrices[0].policies if proj.matrices
            else result.matrices[0].policies
        )

        mismatches = [m.decision_maker for m in result.matrices
                      if m.policies != ref_policies]
        if mismatches:
            msg = ("The following sheets have a different policy list "
                   "and cannot be imported:\n\n"
                   + "\n".join("  - " + dm for dm in mismatches)
                   + "\n\nAll matrices must share the same policy list.")
            messagebox.showerror("Policy List Mismatch", msg, parent=self.root)
            return

        added = 0
        for matrix in result.matrices:
            if matrix.policies == ref_policies:
                proj.matrices.append(matrix)
                self._create_dm_tab(proj, matrix)
                added += 1

        if result.warnings:
            msg = (str(added) + " matrix/matrices imported.\n\n"
                   "The following sheets have blank cells:\n\n"
                   + "\n".join(result.warnings))
            messagebox.showwarning("Import Complete with Warnings",
                                   msg, parent=self.root)
        else:
            messagebox.showinfo("Import Complete",
                                str(added) + " matrix/matrices imported successfully.",
                                parent=self.root)

        self._set_status(f"Imported {added} matrix/matrices from: {path}")
        self._update_topbar()

    def _import_excel_new_project(self):
        path = self._ask_excel_path()
        if not path:
            return

        result = self._load_import_result(path)
        if result is None:
            return

        if not result.matrices:
            messagebox.showwarning("Nothing Imported",
                                   "No valid matrices were found in the workbook.",
                                   parent=self.root)
            return

        project_name = self._unique_project_name(Path(path).stem)
        proj = self._add_project(project_name)
        for matrix in result.matrices:
            proj.matrices.append(matrix)
            self._create_dm_tab(proj, matrix)

        if result.warnings:
            msg = (str(len(result.matrices)) + " matrix/matrices imported.\n\n"
                   "The following sheets have blank cells:\n\n"
                   + "\n".join(result.warnings))
            messagebox.showwarning("Import Complete with Warnings",
                                   msg, parent=self.root)
        else:
            messagebox.showinfo("Import Complete",
                                str(len(result.matrices)) + " matrix/matrices imported successfully.",
                                parent=self.root)

        self._proj_nb.select(proj.frame)
        self._refresh_sidebar_projects()
        self._set_status(f'Project "{proj.name}" created from import: {path}')
        self._update_topbar()

    # ==================================================================
    # Export (operates on current project)
    # ==================================================================

    def _export_excel(self):
        proj = self._current_project()
        if not proj:
            messagebox.showinfo("No Project", "Create a project first.",
                                parent=self.root)
            return
        if not proj.matrices:
            messagebox.showinfo("Nothing to Export",
                                "Add at least one matrix before exporting.",
                                parent=self.root)
            return

        initial = f"{proj.name}_coherence.xlsx".replace(" ", "_")
        path = filedialog.asksaveasfilename(
            parent=self.root,
            title=f"Export project: {proj.name}",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
            initialfile=initial,
        )
        if not path:
            return

        try:
            _export_to_excel(proj.matrices, proj.agg_results, path)
            self._set_status(f"Exported to: {path}")
            messagebox.showinfo("Export Complete",
                                f"Saved to:\n{path}", parent=self.root)
        except Exception as exc:
            messagebox.showerror("Export Failed", str(exc), parent=self.root)

    # ==================================================================
    # Status
    # ==================================================================

    def _set_status(self, message: str):
        self._status_var.set(message)


# =============================================================================
# Excel export helper
# =============================================================================

def _export_to_excel(matrices, agg_results, path: str):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        raise ImportError("openpyxl is required.\nInstall: pip install openpyxl")

    from constants import RATING_COLORS, RATING_TEXT_COLORS
    from coherence_scores_tab import compute_scores
    from range_of_influence_tab import compute_entropy
    from network_tab import compute_centrality

    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def auto_width(ws):
        for col in ws.columns:
            letter  = col[0].column_letter
            max_len = max((len(str(c.value)) for c in col if c.value), default=0)
            ws.column_dimensions[letter].width = max(max_len + 4, 14)

    def hcell(ws, row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(bold=True, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor="2C4A6E")
        c.alignment = Alignment(horizontal="center")
        c.border    = border
        return c

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # DM matrices
    for matrix in matrices:
        ws    = wb.create_sheet(title=matrix.decision_maker[:31])
        n     = len(matrix.policies)
        codes = matrix.codes
        ws.cell(row=1, column=1, value="Policy Index").font = Font(bold=True)
        for k, (code, name) in enumerate(zip(codes, matrix.policies)):
            ws.cell(row=1, column=k+2, value=f"{code}: {name}")
        ws.cell(row=3, column=1, value="Influencing / Influenced")
        ws.cell(row=3, column=1).font = Font(bold=True, italic=True)
        for j, code in enumerate(codes):
            hcell(ws, 3, j+2, code)
        for i, rc in enumerate(codes):
            rh = ws.cell(row=i+4, column=1, value=rc)
            rh.font = Font(bold=True, color="FFFFFF")
            rh.fill = PatternFill("solid", fgColor="2C4A6E")
            rh.alignment = Alignment(horizontal="center")
            rh.border = border
            for j in range(n):
                value = matrix.get_rating(i, j)
                cell  = ws.cell(row=i+4, column=j+2, value=value)
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
                if value in RATING_COLORS:
                    cell.fill = PatternFill("solid",
                        fgColor=RATING_COLORS[value].lstrip("#").upper())
                    cell.font = Font(
                        color=RATING_TEXT_COLORS[value].lstrip("#").upper())
        auto_width(ws)

    # Analysis results
    for idx, result in enumerate(agg_results):
        ml     = {"average":"Avg","majority":"Majority","weighted":"Weighted"
                  }.get(result.method, result.method.title())
        suffix = f" {idx+1}" if len(agg_results) > 1 else ""
        pfx    = f"{ml}{suffix}"
        n, codes = result.n, result.codes

        # Aggregated matrix
        ws = wb.create_sheet(title=f"Aggregated ({pfx})"[:31])
        ws.cell(row=1, column=1, value="Policy Index").font = Font(bold=True)
        for k, (code, name) in enumerate(zip(codes, result.policies)):
            ws.cell(row=1, column=k+2, value=f"{code}: {name}")
        ws.cell(row=3, column=1, value="Influencing / Influenced")
        ws.cell(row=3, column=1).font = Font(bold=True, italic=True)
        for j, code in enumerate(codes): hcell(ws, 3, j+2, code)
        for i, rc in enumerate(codes):
            rh = ws.cell(row=i+4, column=1, value=rc)
            rh.font = Font(bold=True, color="FFFFFF")
            rh.fill = PatternFill("solid", fgColor="2C4A6E")
            rh.alignment = Alignment(horizontal="center")
            rh.border = border
            for j in range(n):
                s = result.scores.get((i,j), 0.0) or 0.0
                cell = ws.cell(row=i+4, column=j+2, value=s)
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
                cell.number_format = "0.00"
        auto_width(ws)

        # Coherence scores
        ws = wb.create_sheet(title=f"Coh Scores ({pfx})"[:31])
        for c, h in enumerate(["Code","Policy","OI","II","WOI","WII"], 1):
            hcell(ws, 1, c, h)
        for r, row in enumerate(compute_scores(result), 2):
            for c, v in enumerate([row["code"], row["policy"], row["oi"],
                                    row["ii"], row["woi"], row["wii"]], 1):
                ws.cell(row=r, column=c, value=v)
        auto_width(ws)

        # Range of influence
        ws = wb.create_sheet(title=f"Range Influence ({pfx})"[:31])
        for c, h in enumerate(["Code","Policy","Entropy","Category"], 1):
            hcell(ws, 1, c, h)
        for r, row in enumerate(compute_entropy(result), 2):
            for c, v in enumerate([row["code"], row["policy"],
                                    row["entropy"], row["category"]], 1):
                ws.cell(row=r, column=c, value=v)
        auto_width(ws)

        # Network centrality
        ws = wb.create_sheet(title=f"Centrality ({pfx})"[:31])
        for c, h in enumerate(["Code","Policy","Betweenness","Closeness"], 1):
            hcell(ws, 1, c, h)
        for r, row in enumerate(
                sorted(compute_centrality(result),
                       key=lambda x: x["betweenness"], reverse=True), 2):
            for c, v in enumerate([row["code"], row["policy"],
                                    row["betweenness"], row["closeness"]], 1):
                ws.cell(row=r, column=c, value=v)
        auto_width(ws)

    wb.save(path)
