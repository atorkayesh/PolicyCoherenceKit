# =============================================================================
# Policy Coherence Kit -- app.py
# Multi-project workspace. Each project lives in its own top-level notebook tab
# and has completely independent state: matrices, analysis results, etc.
# =============================================================================

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from typing import List, Optional
from dataclasses import dataclass, field

from .models import PolicyMatrix, make_empty_matrix
from .matrix_widget import MatrixWidget
from .dialogs import NewMatrixDialog, _SimpleInputDialog, style_button
from .aggregator import (
    check_completeness,
    aggregate_average, aggregate_majority, aggregate_weighted, resolve_ties,
    AggregationResult,
)
from .aggregation_dialog import (
    AggregationMethodDialog, WeightDialog, TieResolutionDialog,
)
from .aggregation_tab import AggregationTab
from .coherence_scores_tab import CoherenceScoresTab
from .range_of_influence_tab import RangeOfInfluenceTab
from .pca_tab import PCATab
from .network_tab import NetworkTab
from .llm_tab import LLMInterpretationTab
from .importer import import_matrices_from_excel
from .constants import (
    APP_TITLE, WINDOW_MIN_WIDTH, WINDOW_MIN_HEIGHT,
    FONT_FAMILY, FONT_SIZE_SMALL, FONT_SIZE_NORMAL,
    FONT_SIZE_HEADER, FONT_SIZE_TITLE,
    COLOR_BG, COLOR_PANEL, COLOR_ACCENT, COLOR_ACCENT2,
    COLOR_TEXT, COLOR_TEXT_LIGHT, COLOR_BORDER,
    COLOR_TAB_BG, COLOR_BUTTON, COLOR_BUTTON_FG,
)


# =============================================================================
# Project data container
# =============================================================================

@dataclass
class Project:
    """All state belonging to one project."""
    name:        str
    matrices:    List[PolicyMatrix] = field(default_factory=list)
    agg_results: List[AggregationResult] = field(default_factory=list)
    agg_tab_ids: List[str] = field(default_factory=list)   # inner notebook tab ids
    notebook:    Optional[ttk.Notebook] = field(default=None, repr=False)
    frame:       Optional[tk.Frame]     = field(default=None, repr=False)


# =============================================================================
# Project name dialog (shown on launch + new project)
# =============================================================================

class _ProjectNameDialog(tk.Toplevel):
    """Simple modal to collect a project name."""

    def __init__(self, parent, title="New Project", existing_names=None):
        super().__init__(parent)
        self.transient(parent)
        self.title(title)
        self.configure(bg=COLOR_BG)
        self.resizable(False, False)
        self.grab_set()
        self._existing = existing_names or []
        self.result: Optional[str] = None
        self._build()
        self.update_idletasks()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        w, h   = self.winfo_reqwidth(), self.winfo_reqheight()
        self.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

    def _build(self):
        tk.Label(
            self, text="Project Name",
            font=(FONT_FAMILY, FONT_SIZE_TITLE, "bold"),
            bg=COLOR_BG, fg=COLOR_ACCENT,
        ).pack(anchor="w", padx=24, pady=(20, 4))
        tk.Label(
            self,
            text="Give this project a unique name.",
            font=(FONT_FAMILY, FONT_SIZE_NORMAL, "italic"),
            bg=COLOR_BG, fg=COLOR_TEXT_LIGHT,
        ).pack(anchor="w", padx=24, pady=(0, 10))
        ttk.Separator(self).pack(fill="x", padx=24, pady=4)
        self._var = tk.StringVar()
        self._entry = tk.Entry(
            self, textvariable=self._var,
            font=(FONT_FAMILY, FONT_SIZE_NORMAL),
            bg="#ffffff", fg=COLOR_TEXT,
            insertbackground=COLOR_ACCENT,
            relief="solid", bd=1, width=36,
        )
        self._entry.pack(padx=24, pady=(8, 16))
        self._entry.focus_set()
        ttk.Separator(self).pack(fill="x", padx=24, pady=4)
        btn_row = tk.Frame(self, bg=COLOR_BG)
        btn_row.pack(anchor="e", padx=24, pady=(6, 18))
        ok = tk.Button(btn_row, text="Create Project", command=self._on_ok)
        style_button(ok)
        ok.pack()
        self.bind("<Return>", lambda e: self._on_ok())
        self.bind("<Escape>", lambda e: self.destroy())

    def _on_ok(self):
        name = self._var.get().strip()
        if not name:
            messagebox.showwarning("Name Required",
                                   "Please enter a project name.", parent=self)
            return
        if name in self._existing:
            messagebox.showwarning("Duplicate Name",
                                   f'A project named "{name}" already exists.',
                                   parent=self)
            return
        self.result = name
        self.destroy()


# =============================================================================
# Main application
# =============================================================================

class PolicyCoherenceApp:
    """
    Multi-project workspace controller.

    Top-level notebook  : one tab per project
    Inner notebook      : DM matrices + analysis tabs for that project
    """

    def __init__(self, root: tk.Tk):
        self.root     = root
        self.projects: List[Project] = []

        self._configure_root()
        self._apply_styles()
        self._build_ui()


    # ==================================================================
    # Setup
    # ==================================================================

    def _configure_root(self):
        self.root.title(APP_TITLE)
        self.root.minsize(WINDOW_MIN_WIDTH, WINDOW_MIN_HEIGHT)
        self.root.configure(bg=COLOR_BG)
        self.root.update_idletasks()
        w, h = 1200, 780
        sw, sh = self.root.winfo_screenwidth(), self.root.winfo_screenheight()
        self.root.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

    def _apply_styles(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TNotebook", background=COLOR_BG,
                        borderwidth=0, tabmargins=[0, 0, 0, 0])
        style.configure("TNotebook.Tab", background=COLOR_TAB_BG,
                        foreground=COLOR_TEXT, padding=[14, 6],
                        font=(FONT_FAMILY, FONT_SIZE_NORMAL))
        style.map("TNotebook.Tab",
                  background=[("selected", COLOR_ACCENT)],
                  foreground=[("selected", "#ffffff")],
                  expand=[("selected", [1, 1, 1, 0])])
        style.configure("TScrollbar", background=COLOR_PANEL,
                        troughcolor=COLOR_BG, borderwidth=0, arrowsize=12)
        style.configure("TCombobox", fieldbackground="#ffffff",
                        background="#ffffff",
                        selectbackground=COLOR_ACCENT,
                        selectforeground="#ffffff")
        style.configure("TSeparator", background=COLOR_BORDER)

        # Inner notebook style (slightly different tab colour)
        style.configure("Inner.TNotebook", background=COLOR_BG,
                        borderwidth=0, tabmargins=[0, 0, 0, 0])
        style.configure("Inner.TNotebook.Tab",
                        background="#c8c4bc",
                        foreground=COLOR_TEXT,
                        padding=[10, 4],
                        font=(FONT_FAMILY, FONT_SIZE_SMALL))
        style.map("Inner.TNotebook.Tab",
                  background=[("selected", COLOR_ACCENT2)],
                  foreground=[("selected", "#ffffff")],
                  expand=[("selected", [1, 1, 1, 0])])

    # ==================================================================
    # Build UI
    # ==================================================================

    def _build_ui(self):
        self._build_header()
        self._build_toolbar()
        tk.Frame(self.root, bg=COLOR_ACCENT2, height=2).pack(fill="x")
        self._build_project_notebook()
        self._build_statusbar()

    def _build_header(self):
        header = tk.Frame(self.root, bg=COLOR_ACCENT, height=56)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(header, text="Policy Coherence Kit",
                 font=(FONT_FAMILY, FONT_SIZE_TITLE, "bold"),
                 bg=COLOR_ACCENT, fg="#ffffff").pack(side="left", padx=(20,8), pady=10)
        tk.Label(header,
                 text="Assess interactions between policies across decision-makers",
                 font=(FONT_FAMILY, FONT_SIZE_NORMAL, "italic"),
                 bg=COLOR_ACCENT, fg="#c8d8ea").pack(side="left", pady=10)

    def _build_toolbar(self):
        toolbar = tk.Frame(self.root, bg=COLOR_PANEL, pady=8)
        toolbar.pack(fill="x")

        # Left side
        new_proj_btn = tk.Button(toolbar, text="  + New Project",
                                 command=self._new_project)
        style_button(new_proj_btn)
        new_proj_btn.pack(side="left", padx=(16, 4))

        delete_proj_btn = tk.Button(toolbar, text="Delete Project",
                                    command=self._delete_project)
        delete_proj_btn.config(bg=COLOR_PANEL, fg="#b71c1c",
                               activebackground=COLOR_PANEL,
                               activeforeground="#b71c1c",
                               relief="flat", padx=10, pady=5,
                               font=(FONT_FAMILY, FONT_SIZE_NORMAL),
                               cursor="hand2")
        delete_proj_btn.pack(side="left", padx=(0, 6))

        tk.Frame(toolbar, bg=COLOR_BORDER, width=1).pack(
            side="left", fill="y", padx=8, pady=4)

        add_dm_btn = tk.Button(toolbar, text="  + Add Decision-Maker",
                               command=self._add_matrix)
        add_dm_btn.config(bg=COLOR_PANEL, fg=COLOR_ACCENT,
                          activebackground=COLOR_PANEL, activeforeground=COLOR_ACCENT,
                          relief="flat", padx=10, pady=5,
                          font=(FONT_FAMILY, FONT_SIZE_NORMAL, "bold"),
                          cursor="hand2")
        add_dm_btn.pack(side="left", padx=4)

        for label, cmd, fg in [
            ("Rename Tab",    self._rename_current_tab,  COLOR_ACCENT),
            ("Remove Tab",    self._remove_current_tab,  "#b71c1c"),
        ]:
            btn = tk.Button(toolbar, text=label, command=cmd)
            btn.config(bg=COLOR_PANEL, fg=fg,
                       activebackground=COLOR_PANEL, activeforeground=fg,
                       relief="flat", padx=10, pady=5,
                       font=(FONT_FAMILY, FONT_SIZE_NORMAL), cursor="hand2")
            btn.pack(side="left", padx=4)

        tk.Frame(toolbar, bg=COLOR_BORDER, width=1).pack(
            side="left", fill="y", padx=8, pady=4)

        agg_btn = tk.Button(toolbar, text="  Run Analysis",
                            command=self._run_aggregation)
        agg_btn.config(bg=COLOR_PANEL, fg="#1a6e3c",
                       activebackground=COLOR_PANEL, activeforeground="#1a6e3c",
                       relief="flat", padx=10, pady=5,
                       font=(FONT_FAMILY, FONT_SIZE_NORMAL, "bold"),
                       cursor="hand2")
        agg_btn.pack(side="left", padx=4)

        # Right side
        export_btn = tk.Button(toolbar, text="  Export to Excel",
                               command=self._export_excel)
        export_btn.config(bg=COLOR_PANEL, fg=COLOR_ACCENT2,
                          activebackground=COLOR_PANEL, activeforeground=COLOR_ACCENT,
                          relief="flat", padx=10, pady=5,
                          font=(FONT_FAMILY, FONT_SIZE_NORMAL), cursor="hand2")
        export_btn.pack(side="right", padx=(4, 16))

        import_btn = tk.Button(toolbar, text="  Import Excel",
                               command=self._import_excel)
        import_btn.config(bg=COLOR_PANEL, fg=COLOR_ACCENT2,
                          activebackground=COLOR_PANEL, activeforeground=COLOR_ACCENT,
                          relief="flat", padx=10, pady=5,
                          font=(FONT_FAMILY, FONT_SIZE_NORMAL), cursor="hand2")
        import_btn.pack(side="right", padx=(0, 4))

    def _build_project_notebook(self):
        self._proj_nb_container = tk.Frame(self.root, bg=COLOR_BG)
        self._proj_nb_container.pack(fill="both", expand=True)
        self._proj_nb = ttk.Notebook(self._proj_nb_container)
        self._proj_nb.pack(fill="both", expand=True)

        self._empty_label = tk.Label(
            self._proj_nb_container,
            text='Click  "+ New Project"  to get started.',
            font=(FONT_FAMILY, FONT_SIZE_HEADER),
            bg=COLOR_BG, fg=COLOR_TEXT_LIGHT, justify="center",
        )
        self._empty_label.place(relx=0.5, rely=0.5, anchor="center")

    def _build_statusbar(self):
        self._status_var = tk.StringVar(value="Ready")
        tk.Label(self.root, textvariable=self._status_var,
                 font=(FONT_FAMILY, FONT_SIZE_SMALL),
                 bg=COLOR_PANEL, fg=COLOR_TEXT_LIGHT,
                 anchor="w", padx=12, pady=4).pack(fill="x", side="bottom")

    # ==================================================================
    # Project management
    # ==================================================================

    def _new_project(self):
        existing = [p.name for p in self.projects]
        dlg = _ProjectNameDialog(self.root, existing_names=existing)
        self.root.wait_window(dlg)
        if dlg.result:
            self._add_project(dlg.result)

    def _add_project(self, name: str):
        """Create a new Project and its top-level notebook tab."""
        self._empty_label.place_forget()

        proj = Project(name=name)

        # Outer frame inside the project notebook
        outer = tk.Frame(self._proj_nb, bg=COLOR_BG)
        self._proj_nb.add(outer, text=f"  {name}  ")
        self._proj_nb.select(outer)
        proj.frame = outer

        # Project header bar
        ph = tk.Frame(outer, bg=COLOR_ACCENT, height=32)
        ph.pack(fill="x")
        ph.pack_propagate(False)
        tk.Label(ph, text=f"Project:  {name}",
                 font=(FONT_FAMILY, FONT_SIZE_NORMAL, "bold"),
                 bg=COLOR_ACCENT, fg="#ffffff").pack(side="left", padx=12, pady=6)

        # Inner notebook for DM tabs + analysis tabs
        inner_nb = ttk.Notebook(outer, style="Inner.TNotebook")
        inner_nb.pack(fill="both", expand=True)
        proj.notebook = inner_nb

        # Empty state inside project
        empty = tk.Label(
            outer,
            text='Click  "+ Add Decision-Maker"  to add the first matrix.',
            font=(FONT_FAMILY, FONT_SIZE_HEADER),
            bg=COLOR_BG, fg=COLOR_TEXT_LIGHT,
        )
        empty.place(relx=0.5, rely=0.55, anchor="center")
        proj._empty_label = empty

        self.projects.append(proj)
        self._set_status(f'Project "{name}" created.')

    def _current_project(self) -> Optional[Project]:
        """Return the currently selected project, or None."""
        if not self._proj_nb.tabs():
            return None
        try:
            idx = self._proj_nb.index("current")
            return self.projects[idx]
        except (tk.TclError, IndexError):
            return None

    def _rename_project(self):
        proj = self._current_project()
        if not proj:
            messagebox.showinfo("No Project", "No project is open.", parent=self.root)
            return
        existing = [p.name for p in self.projects if p is not proj]
        dlg = _SimpleInputDialog(self.root, "Rename Project",
                                 "New project name:", proj.name)
        self.root.wait_window(dlg)
        if dlg.result and dlg.result != proj.name:
            if dlg.result in [p.name for p in self.projects]:
                messagebox.showwarning("Duplicate Name",
                                       f'A project named "{dlg.result}" already exists.',
                                       parent=self.root)
                return
            idx = self._proj_nb.index("current")
            proj.name = dlg.result
            self._proj_nb.tab(idx, text=f"  {dlg.result}  ")
            # Update project header bar
            for widget in proj.frame.winfo_children():
                if isinstance(widget, tk.Frame) and widget.cget("bg") == COLOR_ACCENT:
                    for lbl in widget.winfo_children():
                        if isinstance(lbl, tk.Label):
                            lbl.config(text=f"Project:  {dlg.result}")
            self._set_status(f'Project renamed to "{dlg.result}".')

    def _delete_project(self):
        proj = self._current_project()
        if not proj:
            messagebox.showinfo("No Project", "No project is open.",
                                parent=self.root)
            return
        if not messagebox.askyesno(
            "Delete Project",
            f'Permanently delete project "{proj.name}" and all its data?\n'
            "This cannot be undone.",
            parent=self.root,
        ):
            return
        idx = self._proj_nb.index("current")
        self._proj_nb.forget(idx)
        self.projects.pop(idx)
        if not self.projects:
            self._empty_label.place(relx=0.5, rely=0.5, anchor="center")
        self._set_status(f'Project "{proj.name}" deleted.')

    # ==================================================================
    # Decision-maker tab management (operates on current project)
    # ==================================================================

    def _add_matrix(self):
        proj = self._current_project()
        if not proj:
            messagebox.showinfo("No Project",
                                "Create a project first.", parent=self.root)
            return

        if proj.matrices:
            # Reuse existing policy list
            dlg = _SimpleInputDialog(self.root, "Add Decision-Maker",
                                     "Decision-Maker Name:")
            self.root.wait_window(dlg)
            if not dlg.result:
                return
            dm_name  = dlg.result
            policies = proj.matrices[0].policies
        else:
            dlg = NewMatrixDialog(self.root)
            self.root.wait_window(dlg)
            if dlg.result is None:
                return
            dm_name, policies = dlg.result

        matrix = make_empty_matrix(dm_name, policies)
        proj.matrices.append(matrix)
        self._create_dm_tab(proj, matrix)
        self._set_status(f'"{dm_name}" added to project "{proj.name}".')

    def _create_dm_tab(self, proj: Project, matrix: PolicyMatrix):
        """Build one DM matrix tab inside the project's inner notebook."""
        proj._empty_label.place_forget()

        tab = tk.Frame(proj.notebook, bg=COLOR_BG)
        proj.notebook.add(tab, text=f"  {matrix.decision_maker}  ")
        proj.notebook.select(tab)

        # Info bar
        info = tk.Frame(tab, bg=COLOR_PANEL, pady=6)
        info.pack(fill="x")
        tk.Label(info, text=f"Decision-Maker:  {matrix.decision_maker}",
                 font=(FONT_FAMILY, FONT_SIZE_HEADER, "bold"),
                 bg=COLOR_PANEL, fg=COLOR_ACCENT).pack(side="left", padx=16)
        tk.Label(info, text=f"{len(matrix.policies)} policies  |  "
                            f"{matrix.total_cells()} cells to fill",
                 font=(FONT_FAMILY, FONT_SIZE_NORMAL, "italic"),
                 bg=COLOR_PANEL, fg=COLOR_TEXT_LIGHT).pack(side="left", padx=6)

        tk.Frame(tab, bg=COLOR_BORDER, height=1).pack(fill="x", pady=4)

        def on_change(r, c, v):
            filled = matrix.filled_count()
            total  = matrix.total_cells()
            self._set_status(
                f'"{matrix.decision_maker}"  --  {filled}/{total} cells filled  '
                f'|  {matrix.codes[r]} -> {matrix.codes[c]}: {v}'
            )

        mw = MatrixWidget(tab, matrix, on_change=on_change)
        mw.pack(fill="both", expand=True, padx=8, pady=8)

    def _current_inner_index(self, proj: Project) -> int:
        if not proj.notebook.tabs():
            return -1
        try:
            return proj.notebook.index("current")
        except tk.TclError:
            return -1

    def _rename_current_tab(self):
        proj = self._current_project()
        if not proj:
            return
        idx = self._current_inner_index(proj)
        if idx < 0:
            return
        # Only DM tabs (before agg tabs) can be renamed
        n_dm = len(proj.matrices)
        if idx >= n_dm:
            messagebox.showinfo("Cannot Rename",
                                "Analysis result tabs cannot be renamed.",
                                parent=self.root)
            return
        matrix = proj.matrices[idx]
        dlg = _SimpleInputDialog(self.root, "Rename Tab",
                                 "New decision-maker name:", matrix.decision_maker)
        self.root.wait_window(dlg)
        if dlg.result:
            matrix.decision_maker = dlg.result
            proj.notebook.tab(idx, text=f"  {dlg.result}  ")
            self._set_status(f'Renamed to "{dlg.result}".')

    def _remove_current_tab(self):
        proj = self._current_project()
        if not proj:
            return
        idx = self._current_inner_index(proj)
        if idx < 0:
            return
        n_dm = len(proj.matrices)

        if idx >= n_dm:
            # Analysis result tab
            if not messagebox.askyesno("Remove Tab",
                                       "Remove this analysis result tab?",
                                       parent=self.root):
                return
            tab_id = proj.notebook.tabs()[idx]
            proj.notebook.forget(idx)
            if tab_id in proj.agg_tab_ids:
                proj.agg_tab_ids.remove(tab_id)
            self._set_status("Analysis tab removed.")
        else:
            # DM tab
            dm = proj.matrices[idx].decision_maker
            if not messagebox.askyesno("Remove Matrix",
                                       f'Remove matrix for "{dm}"?',
                                       parent=self.root):
                return
            proj.notebook.forget(idx)
            proj.matrices.pop(idx)
            if not proj.matrices:
                proj._empty_label.place(relx=0.5, rely=0.55, anchor="center")
            self._set_status(f'Removed "{dm}".')

    # ==================================================================
    # Aggregation / Analysis (operates on current project)
    # ==================================================================

    def _run_aggregation(self):
        proj = self._current_project()
        if not proj:
            messagebox.showinfo("No Project", "Create a project first.",
                                parent=self.root)
            return

        if len(proj.matrices) < 2:
            messagebox.showwarning("Not Enough Matrices",
                                   "You need at least 2 decision-maker matrices.",
                                   parent=self.root)
            return

        incomplete = check_completeness(proj.matrices)
        if incomplete:
            lines = []
            for dm_name, blanks in incomplete:
                cells = ", ".join(f"{r}->{c}" for r, c in blanks[:10])
                suffix = f" (+{len(blanks)-10} more)" if len(blanks) > 10 else ""
                lines.append(f"  {dm_name}: {cells}{suffix}")
            messagebox.showerror("Incomplete Matrices",
                                 "Please fill all cells before aggregating.\n\n"
                                 + "\n".join(lines), parent=self.root)
            return

        n_dms = len(proj.matrices)
        if n_dms < 3:
            method = "average"
        else:
            dlg = AggregationMethodDialog(
                self.root, [m.decision_maker for m in proj.matrices])
            self.root.wait_window(dlg)
            if dlg.result is None:
                return
            method = dlg.result

        weights = None
        if method == "weighted":
            dlg_w = WeightDialog(
                self.root, [m.decision_maker for m in proj.matrices])
            self.root.wait_window(dlg_w)
            if dlg_w.result is None:
                return
            weights = dlg_w.result

        if method == "average":
            result = aggregate_average(proj.matrices)
        elif method == "majority":
            result = aggregate_majority(proj.matrices)
        elif method == "weighted":
            result = aggregate_weighted(proj.matrices, weights)
        else:
            return

        if result.ties:
            dlg_t = TieResolutionDialog(self.root, result.ties)
            self.root.wait_window(dlg_t)
            if dlg_t.result is None:
                return
            resolve_ties(result)

        self._create_analysis_tabs(proj, result)
        self._set_status(
            f'Analysis complete for "{proj.name}"  |  method: {method}  '
            f'|  {n_dms} decision-makers'
        )

    def _create_analysis_tabs(self, proj: Project, result: AggregationResult):
        """Add all six analysis tabs to the project's inner notebook."""
        method_label = {
            "average":  "Average",
            "majority": "Majority",
            "weighted": "Weighted",
        }.get(result.method, result.method.title())

        proj.agg_results.append(result)

        tabs = [
            (f"  Aggregated ({method_label})  ",       AggregationTab),
            (f"  Coherence Scores ({method_label})  ", CoherenceScoresTab),
            (f"  Range of Influence ({method_label})  ", RangeOfInfluenceTab),
            (f"  PCA ({method_label})  ",              PCATab),
            (f"  Network Analysis ({method_label})  ", NetworkTab),
            (f"  LLM Interpretation ({method_label})  ", LLMInterpretationTab),
        ]

        first = True
        for title, TabClass in tabs:
            widget = TabClass(proj.notebook, result)
            proj.notebook.add(widget, text=title)
            tab_id = proj.notebook.tabs()[-1]
            proj.agg_tab_ids.append(tab_id)
            if first:
                proj.notebook.select(widget)
                first = False

    # ==================================================================
    # Import (operates on current project)
    # ==================================================================

    def _import_excel(self):
        proj = self._current_project()
        if not proj:
            messagebox.showinfo("No Project", "Create a project first.",
                                parent=self.root)
            return

        path = filedialog.askopenfilename(
            parent=self.root,
            title="Import matrices from Excel workbook",
            filetypes=[("Excel Workbook", "*.xlsx *.xls")],
        )
        if not path:
            return

        try:
            result = import_matrices_from_excel(path)
        except (ValueError, ImportError) as exc:
            messagebox.showerror("Import Error", str(exc), parent=self.root)
            return

        if not result.matrices:
            messagebox.showwarning("Nothing Imported",
                                   "No valid matrices were found in the workbook.",
                                   parent=self.root)
            return

        ref_policies = (
            proj.matrices[0].policies if proj.matrices
            else result.matrices[0].policies
        )

        mismatches = [m.decision_maker for m in result.matrices
                      if m.policies != ref_policies]
        if mismatches:
            msg = ("The following sheets have a different policy list "
                   "and cannot be imported:\n\n"
                   + "\n".join("  - " + dm for dm in mismatches)
                   + "\n\nAll matrices must share the same policy list.")
            messagebox.showerror("Policy List Mismatch", msg, parent=self.root)
            return

        added = 0
        for matrix in result.matrices:
            if matrix.policies == ref_policies:
                proj.matrices.append(matrix)
                self._create_dm_tab(proj, matrix)
                added += 1

        if result.warnings:
            msg = (str(added) + " matrix/matrices imported.\n\n"
                   "The following sheets have blank cells:\n\n"
                   + "\n".join(result.warnings))
            messagebox.showwarning("Import Complete with Warnings",
                                   msg, parent=self.root)
        else:
            messagebox.showinfo("Import Complete",
                                str(added) + " matrix/matrices imported successfully.",
                                parent=self.root)

        self._set_status(f"Imported {added} matrix/matrices from: {path}")

    # ==================================================================
    # Export (operates on current project)
    # ==================================================================

    def _export_excel(self):
        proj = self._current_project()
        if not proj:
            messagebox.showinfo("No Project", "Create a project first.",
                                parent=self.root)
            return
        if not proj.matrices:
            messagebox.showinfo("Nothing to Export",
                                "Add at least one matrix before exporting.",
                                parent=self.root)
            return

        initial = f"{proj.name}_coherence.xlsx".replace(" ", "_")
        path = filedialog.asksaveasfilename(
            parent=self.root,
            title=f"Export project: {proj.name}",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
            initialfile=initial,
        )
        if not path:
            return

        try:
            _export_to_excel(proj.matrices, proj.agg_results, path)
            self._set_status(f"Exported to: {path}")
            messagebox.showinfo("Export Complete",
                                f"Saved to:\n{path}", parent=self.root)
        except Exception as exc:
            messagebox.showerror("Export Failed", str(exc), parent=self.root)

    # ==================================================================
    # Status
    # ==================================================================

    def _set_status(self, message: str):
        self._status_var.set(message)


# =============================================================================
# Excel export helper
# =============================================================================

def _export_to_excel(matrices, agg_results, path: str):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        raise ImportError("openpyxl is required.\nInstall: pip install openpyxl")

    from .constants import RATING_COLORS, RATING_TEXT_COLORS
    from .coherence_scores_tab import compute_scores
    from .range_of_influence_tab import compute_entropy
    from .network_tab import compute_centrality

    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def auto_width(ws):
        for col in ws.columns:
            letter  = col[0].column_letter
            max_len = max((len(str(c.value)) for c in col if c.value), default=0)
            ws.column_dimensions[letter].width = max(max_len + 4, 14)

    def hcell(ws, row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(bold=True, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor="2C4A6E")
        c.alignment = Alignment(horizontal="center")
        c.border    = border
        return c

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # DM matrices
    for matrix in matrices:
        ws    = wb.create_sheet(title=matrix.decision_maker[:31])
        n     = len(matrix.policies)
        codes = matrix.codes
        ws.cell(row=1, column=1, value="Policy Index").font = Font(bold=True)
        for k, (code, name) in enumerate(zip(codes, matrix.policies)):
            ws.cell(row=1, column=k+2, value=f"{code}: {name}")
        ws.cell(row=3, column=1, value="Influencing / Influenced")
        ws.cell(row=3, column=1).font = Font(bold=True, italic=True)
        for j, code in enumerate(codes):
            hcell(ws, 3, j+2, code)
        for i, rc in enumerate(codes):
            rh = ws.cell(row=i+4, column=1, value=rc)
            rh.font = Font(bold=True, color="FFFFFF")
            rh.fill = PatternFill("solid", fgColor="2C4A6E")
            rh.alignment = Alignment(horizontal="center")
            rh.border = border
            for j in range(n):
                value = matrix.get_rating(i, j)
                cell  = ws.cell(row=i+4, column=j+2, value=value)
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
                if value in RATING_COLORS:
                    cell.fill = PatternFill("solid",
                        fgColor=RATING_COLORS[value].lstrip("#").upper())
                    cell.font = Font(
                        color=RATING_TEXT_COLORS[value].lstrip("#").upper())
        auto_width(ws)

    # Analysis results
    for idx, result in enumerate(agg_results):
        ml     = {"average":"Avg","majority":"Majority","weighted":"Weighted"
                  }.get(result.method, result.method.title())
        suffix = f" {idx+1}" if len(agg_results) > 1 else ""
        pfx    = f"{ml}{suffix}"
        n, codes = result.n, result.codes

        # Aggregated matrix
        ws = wb.create_sheet(title=f"Aggregated ({pfx})"[:31])
        ws.cell(row=1, column=1, value="Policy Index").font = Font(bold=True)
        for k, (code, name) in enumerate(zip(codes, result.policies)):
            ws.cell(row=1, column=k+2, value=f"{code}: {name}")
        ws.cell(row=3, column=1, value="Influencing / Influenced")
        ws.cell(row=3, column=1).font = Font(bold=True, italic=True)
        for j, code in enumerate(codes): hcell(ws, 3, j+2, code)
        for i, rc in enumerate(codes):
            rh = ws.cell(row=i+4, column=1, value=rc)
            rh.font = Font(bold=True, color="FFFFFF")
            rh.fill = PatternFill("solid", fgColor="2C4A6E")
            rh.alignment = Alignment(horizontal="center")
            rh.border = border
            for j in range(n):
                s = result.scores.get((i,j), 0.0) or 0.0
                cell = ws.cell(row=i+4, column=j+2, value=s)
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
                cell.number_format = "0.00"
        auto_width(ws)

        # Coherence scores
        ws = wb.create_sheet(title=f"Coh Scores ({pfx})"[:31])
        for c, h in enumerate(["Code","Policy","OI","II","WOI","WII"], 1):
            hcell(ws, 1, c, h)
        for r, row in enumerate(compute_scores(result), 2):
            for c, v in enumerate([row["code"], row["policy"], row["oi"],
                                    row["ii"], row["woi"], row["wii"]], 1):
                ws.cell(row=r, column=c, value=v)
        auto_width(ws)

        # Range of influence
        ws = wb.create_sheet(title=f"Range Influence ({pfx})"[:31])
        for c, h in enumerate(["Code","Policy","Entropy","Category"], 1):
            hcell(ws, 1, c, h)
        for r, row in enumerate(compute_entropy(result), 2):
            for c, v in enumerate([row["code"], row["policy"],
                                    row["entropy"], row["category"]], 1):
                ws.cell(row=r, column=c, value=v)
        auto_width(ws)

        # Network centrality
        ws = wb.create_sheet(title=f"Centrality ({pfx})"[:31])
        for c, h in enumerate(["Code","Policy","Betweenness","Closeness"], 1):
            hcell(ws, 1, c, h)
        for r, row in enumerate(
                sorted(compute_centrality(result),
                       key=lambda x: x["betweenness"], reverse=True), 2):
            for c, v in enumerate([row["code"], row["policy"],
                                    row["betweenness"], row["closeness"]], 1):
                ws.cell(row=r, column=c, value=v)
        auto_width(ws)

    wb.save(path)
